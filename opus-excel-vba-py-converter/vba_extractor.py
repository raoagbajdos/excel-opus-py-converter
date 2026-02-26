"""
VBA Extractor Module
Extracts VBA code from Excel files (.xlsm, .xls, .xlsb, .xla, .xlam)
"""
import os
import zipfile
import tempfile
import re
from typing import List, Dict, Optional


class VBAExtractor:
    """Extract VBA code from Excel files."""
    
    _CLASS_MODULE = 'Class Module'
    
    # Module type identifiers
    MODULE_TYPES = {
        1: 'Standard Module',
        2: _CLASS_MODULE, 
        3: 'UserForm',
        100: 'Document Module (ThisWorkbook/Sheet)'
    }
    
    def __init__(self, filepath: str):
        """
        Initialize the VBA extractor.
        
        Args:
            filepath: Path to the Excel file
        """
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self.extension = os.path.splitext(filepath)[1].lower()
        
    def extract_all(self) -> List[Dict]:
        """
        Extract all VBA modules from the Excel file.
        
        Returns:
            List of dictionaries containing module information
        """
        # Always try oletools first — it auto-detects the format (OLE or OpenXML)
        modules: List[Dict] = []
        try:
            from oletools.olevba import VBA_Parser
            vba_parser = VBA_Parser(self.filepath)
            if vba_parser.detect_vba_macros():
                for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                    if vba_code and vba_code.strip():
                        module_type = self._detect_module_type(vba_filename, vba_code)
                        modules.append({
                            'name': vba_filename or 'Unknown',
                            'type': module_type,
                            'code': vba_code,
                            'stream_path': stream_path
                        })
            vba_parser.close()
        except ImportError:
            pass
        except Exception as e:
            # oletools couldn't handle the file — fall through to manual methods
            import logging
            logging.getLogger(__name__).debug("oletools failed: %s", e)

        # If no embedded macros found, try extracting VBA stored as text in sheets
        if not modules:
            sheet_modules = self._extract_vba_from_sheet_cells()
            if sheet_modules:
                return sheet_modules

        if modules:
            return modules

        # Fallback: format-specific extraction without oletools
        if self.extension in ['.xlsm', '.xlsb', '.xlam', '.xlsx']:
            return self._extract_from_xlsx_format()
        elif self.extension in ['.xls', '.xla']:
            return self._manual_ole_extraction()
        else:
            raise ValueError(f"Unsupported file format: {self.extension}")
    
    def _extract_vba_from_sheet_cells(self) -> List[Dict]:
        """Extract VBA code stored as text in worksheet cells.

        Some workbooks store VBA source code as plain text in a dedicated
        worksheet (e.g. "VBA_Code", "Macros", "VBA", "Code") rather than
        embedding it in a vbaProject.bin OLE stream.

        The method scans for such sheets, reads all non-empty cells, and
        splits the concatenated text into individual Sub / Function modules.
        """
        # Sheet names that commonly hold VBA source text
        vba_sheet_names = {
            'vba_code', 'vba', 'macros', 'macro_code', 'code',
            'vbacode', 'macro', 'vba_source', 'source_code',
        }

        try:
            import openpyxl

            wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
        except Exception:
            return []

        modules: List[Dict] = []
        try:
            for sheet_name in wb.sheetnames:
                if sheet_name.lower().replace(' ', '_') not in vba_sheet_names:
                    continue

                ws = wb[sheet_name]
                lines: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value is not None:
                            text = str(cell_value).strip()
                            if text:
                                lines.append(text)

                if not lines:
                    continue

                # Join all cell text into one blob and split into modules
                full_text = '\n'.join(lines)
                parsed = self._split_vba_text_into_modules(full_text, sheet_name)
                modules.extend(parsed)
        finally:
            wb.close()

        return modules

    def _split_vba_text_into_modules(
        self, text: str, source_sheet: str
    ) -> List[Dict]:
        """Split a block of VBA text into individual Sub / Function modules."""
        # Regex to find Sub / Function boundaries
        pattern = re.compile(
            r'(?:^|\n)'                                         # start of text or newline
            r'((?:Public\s+|Private\s+)?'                       # optional scope
            r'(?:Sub|Function)\s+'                              # keyword
            r'[A-Za-z_]\w*'                                     # name
            r'\s*\([^)]*\)'                                     # params
            r'(?:\s+As\s+\w+)?'                                 # optional return type
            r'.*?'                                              # body
            r'End\s+(?:Sub|Function))',                          # closing
            re.IGNORECASE | re.DOTALL,
        )
        matches = list(pattern.finditer(text))

        if not matches:
            # No recognisable Sub/Function – return the entire text as one module
            return [{
                'name': f'{source_sheet}_code',
                'type': 'Standard Module',
                'code': text,
                'stream_path': f'Cells/{source_sheet}',
            }]

        modules: List[Dict] = []
        for m in matches:
            code_block = m.group(0).strip()
            # Extract the procedure name for the module name
            name_match = re.search(
                r'(?:Sub|Function)\s+([A-Za-z_]\w*)', code_block, re.IGNORECASE,
            )
            proc_name = name_match.group(1) if name_match else f'Module_{len(modules)+1}'
            mod_type = self._detect_module_type(proc_name, code_block)
            modules.append({
                'name': proc_name,
                'type': mod_type,
                'code': code_block,
                'stream_path': f'Cells/{source_sheet}',
            })

        # If there's preamble text before the first match (Dim statements, comments, etc.)
        preamble = text[:matches[0].start()].strip()
        if preamble and len(preamble) > 20:
            modules.insert(0, {
                'name': f'{source_sheet}_declarations',
                'type': 'Standard Module',
                'code': preamble,
                'stream_path': f'Cells/{source_sheet}',
            })

        return modules

    def _extract_from_xlsx_format(self) -> List[Dict]:
        """
        Extract VBA from modern Excel formats (.xlsm, .xlsb, .xlam).
        These are ZIP-based formats with vbaProject.bin inside.
        """
        modules = []
        
        try:
            with zipfile.ZipFile(self.filepath, 'r') as zf:
                # Look for vbaProject.bin
                vba_files = [f for f in zf.namelist() if 'vbaProject.bin' in f]
                
                if not vba_files:
                    return []
                
                for vba_file in vba_files:
                    vba_content = zf.read(vba_file)
                    extracted = self._parse_vba_project(vba_content)
                    modules.extend(extracted)
                    
        except zipfile.BadZipFile:
            raise ValueError("Invalid or corrupted Excel file")
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
        
        return modules
    
    def _extract_from_xls_format(self) -> List[Dict]:
        """
        Extract VBA from legacy Excel format (.xls, .xla).
        Uses oletools if available, otherwise attempts manual extraction.
        """
        try:
            # Try using oletools (olevba) if available
            from oletools.olevba import VBA_Parser
            
            vba_parser = VBA_Parser(self.filepath)
            modules = []
            
            if vba_parser.detect_vba_macros():
                for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                    if vba_code and vba_code.strip():
                        module_type = self._detect_module_type(vba_filename, vba_code)
                        modules.append({
                            'name': vba_filename or 'Unknown',
                            'type': module_type,
                            'code': vba_code,
                            'stream_path': stream_path
                        })
            
            vba_parser.close()
            return modules
            
        except ImportError:
            # Fallback to manual extraction for .xls files
            return self._manual_ole_extraction()
    
    def _parse_vba_project(self, vba_content: bytes) -> List[Dict]:
        """
        Parse the vbaProject.bin file to extract VBA code.
        
        Args:
            vba_content: Binary content of vbaProject.bin
            
        Returns:
            List of module dictionaries
        """
        modules = []
        
        try:
            # Try using oletools if available
            from oletools.olevba import VBA_Parser
            
            # Create a temporary file
            with tempfile.NamedTemporaryFile(suffix='.bin', delete=False) as tmp:
                tmp.write(vba_content)
                tmp_path = tmp.name
            
            try:
                vba_parser = VBA_Parser(tmp_path)
                
                if vba_parser.detect_vba_macros():
                    for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                        if vba_code and vba_code.strip():
                            module_type = self._detect_module_type(vba_filename, vba_code)
                            modules.append({
                                'name': vba_filename or 'Unknown',
                                'type': module_type,
                                'code': vba_code,
                                'stream_path': stream_path
                            })
                
                vba_parser.close()
            finally:
                os.unlink(tmp_path)
                
        except ImportError:
            # Fallback to manual extraction
            modules = self._manual_vba_extraction(vba_content)
        
        return modules
    
    def _manual_vba_extraction(self, vba_content: bytes) -> List[Dict]:
        """
        Manual extraction of VBA code from binary content.
        This is a fallback when oletools is not available.
        
        Args:
            vba_content: Binary content containing VBA
            
        Returns:
            List of module dictionaries
        """
        content_str = vba_content.decode('latin-1', errors='ignore')
        modules = self._extract_modules_by_attribute(content_str)

        if not modules:
            modules = self._extract_modules_by_keywords(content_str)
        
        return modules

    def _extract_modules_by_attribute(self, content_str: str) -> List[Dict]:
        """Extract modules using 'Attribute VB_Name' markers."""
        modules: List[Dict] = []
        module_pattern = r'Attribute\s+VB_Name\s*=\s*"([^"]+)"'
        matches = list(re.finditer(module_pattern, content_str))
        
        for i, match in enumerate(matches):
            module_name = match.group(1)
            start_pos = match.start()
            end_pos = matches[i + 1].start() if i + 1 < len(matches) else len(content_str)
            
            code_section = content_str[start_pos:end_pos]
            cleaned_code = self._clean_extracted_code(code_section)
            
            if cleaned_code.strip():
                modules.append({
                    'name': module_name,
                    'type': self._detect_module_type(module_name, cleaned_code),
                    'code': cleaned_code,
                    'stream_path': 'manual_extraction'
                })
        return modules

    def _extract_modules_by_keywords(self, content_str: str) -> List[Dict]:
        """Fallback: extract VBA code by searching for common keywords."""
        vba_keywords = [
            'Sub ', 'Function ', 'Private Sub', 'Public Sub',
            'Private Function', 'Public Function', 'Dim ', 'End Sub', 'End Function',
        ]
        for keyword in vba_keywords:
            if keyword in content_str:
                code_start = content_str.find(keyword)
                extracted = content_str[code_start:code_start + 5000]
                cleaned = self._clean_extracted_code(extracted)
                if cleaned.strip():
                    return [{
                        'name': 'ExtractedCode',
                        'type': 'Unknown',
                        'code': cleaned,
                        'stream_path': 'manual_extraction'
                    }]
        return []
    
    def _manual_ole_extraction(self) -> List[Dict]:
        """
        Manual extraction for OLE compound files (.xls).
        """
        modules = []
        
        try:
            import olefile
            
            ole = olefile.OleFileIO(self.filepath)
            
            # Look for VBA storage
            if ole.exists('_VBA_PROJECT_CUR'):
                vba_root = '_VBA_PROJECT_CUR'
            elif ole.exists('Macros'):
                vba_root = 'Macros'
            else:
                # Try to find VBA in any storage
                for stream in ole.listdir():
                    stream_path = '/'.join(stream)
                    if 'VBA' in stream_path.upper():
                        content = ole.openstream(stream).read()
                        extracted = self._manual_vba_extraction(content)
                        modules.extend(extracted)
                
                ole.close()
                return modules
            
            # Extract from VBA storage
            for stream in ole.listdir():
                stream_path = '/'.join(stream)
                if vba_root in stream_path:
                    try:
                        content = ole.openstream(stream).read()
                        extracted = self._manual_vba_extraction(content)
                        modules.extend(extracted)
                    except Exception:
                        pass
            
            ole.close()
            
        except ImportError:
            # If olefile is not available, try raw binary extraction
            with open(self.filepath, 'rb') as f:
                content = f.read()
            modules = self._manual_vba_extraction(content)
        
        return modules
    
    def _detect_module_type(self, module_name: str, code: str) -> str:
        """
        Detect the type of VBA module based on name and content.
        
        Args:
            module_name: Name of the module
            code: VBA code content
            
        Returns:
            Module type string
        """
        name_lower = module_name.lower() if module_name else ''
        
        if 'thisworkbook' in name_lower:
            return 'Document Module (ThisWorkbook)'
        elif 'sheet' in name_lower:
            return 'Document Module (Sheet)'
        elif 'userform' in name_lower or 'frm' in name_lower:
            return 'UserForm'
        elif 'class' in name_lower or 'cls' in name_lower:
            return self._CLASS_MODULE
        
        # Check code content
        if 'Attribute VB_Creatable' in code or 'Attribute VB_Exposed' in code:
            return self._CLASS_MODULE
        elif 'Begin {' in code or 'Begin VB.Form' in code:
            return 'UserForm'
        
        return 'Standard Module'
    
    def _clean_extracted_code(self, code: str) -> str:
        """
        Clean up extracted VBA code.
        
        Args:
            code: Raw extracted code
            
        Returns:
            Cleaned VBA code
        """
        # Remove null bytes and control characters
        cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', code)
        
        # Remove excessive whitespace
        lines = cleaned.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # Skip lines that are just garbage
            if len(line) > 0 and len(line.strip()) > 0:
                # Check if line contains mostly printable characters
                printable_ratio = sum(c.isprintable() or c.isspace() for c in line) / len(line)
                if printable_ratio > 0.8:
                    cleaned_lines.append(line.rstrip())
        
        return '\n'.join(cleaned_lines)


def extract_vba_from_file(filepath: str) -> List[Dict]:
    """
    Convenience function to extract VBA from an Excel file.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        List of module dictionaries
    """
    extractor = VBAExtractor(filepath)
    return extractor.extract_all()
