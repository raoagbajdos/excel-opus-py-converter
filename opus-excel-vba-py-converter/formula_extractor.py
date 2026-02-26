"""
Formula Extractor Module
Extracts and analyzes Excel formulas from workbooks
"""
import logging
import os
import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Set

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class FormulaInfo:
    """Information about an Excel formula."""
    sheet_name: str
    cell_address: str
    formula: str
    formula_type: str  # 'standard', 'array', 'shared'
    dependencies: List[str]  # Referenced cells/ranges
    contains_functions: List[str]  # Excel functions used


class FormulaExtractor:
    """Extract and analyze formulas from Excel workbooks."""
    
    # Common Excel functions that might need special Python conversion
    EXCEL_FUNCTIONS = {
        # Lookup/Reference
        'VLOOKUP', 'HLOOKUP', 'XLOOKUP', 'INDEX', 'MATCH', 'LOOKUP',
        'OFFSET', 'INDIRECT', 'CHOOSE',
        
        # Math/Stats
        'SUM', 'SUMIF', 'SUMIFS', 'AVERAGE', 'AVERAGEIF', 'AVERAGEIFS',
        'COUNT', 'COUNTA', 'COUNTIF', 'COUNTIFS', 'COUNTBLANK',
        'MAX', 'MIN', 'MAXIFS', 'MINIFS', 'MEDIAN', 'MODE', 'STDEV',
        
        # Logical
        'IF', 'IFS', 'AND', 'OR', 'NOT', 'XOR', 'IFERROR', 'IFNA',
        
        # Text
        'CONCATENATE', 'CONCAT', 'TEXTJOIN', 'LEFT', 'RIGHT', 'MID',
        'LEN', 'FIND', 'SEARCH', 'SUBSTITUTE', 'REPLACE', 'TRIM',
        'UPPER', 'LOWER', 'PROPER', 'TEXT',
        
        # Date/Time
        'TODAY', 'NOW', 'DATE', 'TIME', 'YEAR', 'MONTH', 'DAY',
        'HOUR', 'MINUTE', 'SECOND', 'DATEDIF', 'EOMONTH', 'EDATE',
        'WORKDAY', 'NETWORKDAYS',
        
        # Financial
        'PMT', 'IPMT', 'PPMT', 'FV', 'PV', 'NPV', 'IRR', 'XIRR',
        'RATE', 'NPER',
        
        # Database
        'DSUM', 'DAVERAGE', 'DCOUNT', 'DMAX', 'DMIN',
        
        # Array/Modern
        'FILTER', 'SORT', 'SORTBY', 'UNIQUE', 'SEQUENCE', 'RANDARRAY',
    }
    
    def __init__(self, filepath: str):
        """
        Initialize the formula extractor.
        
        Args:
            filepath: Path to the Excel file
        """
        self.filepath = filepath
        self.workbook: Optional[openpyxl.Workbook] = None
        
    def extract_all_formulas(self) -> List[FormulaInfo]:
        """
        Extract all formulas from all sheets in the workbook.
        
        Returns:
            List of FormulaInfo objects
        """
        ext = os.path.splitext(self.filepath)[1].lower()
        if ext == '.xls':
            return self._extract_formulas_xlrd()

        self.workbook = openpyxl.load_workbook(self.filepath, data_only=False)
        all_formulas = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet_formulas = self._extract_sheet_formulas(sheet_name)
            all_formulas.extend(sheet_formulas)
        
        self.workbook.close()
        return all_formulas

    def _extract_formulas_xlrd(self) -> List[FormulaInfo]:
        """Extract formulas from a genuine .xls (BIFF) file using xlrd.

        xlrd 2.x can read .xls files but does NOT expose formula text—only
        cell values.  We can still detect formula cells (cell type XL_CELL_FORMULA
        = 4) and record what we can.
        """
        try:
            import xlrd
        except ImportError:
            logger.warning("xlrd not installed — cannot read .xls formulas")
            return []

        formulas: List[FormulaInfo] = []
        book = xlrd.open_workbook(self.filepath)
        # xlrd 2.x can read .xls but does NOT expose formula text — only values.
        # We simply return an empty list; users should convert to .xlsx for formulas.
        book.release_resources()
        return formulas
    
    def _extract_sheet_formulas(self, sheet_name: str) -> List[FormulaInfo]:
        """
        Extract formulas from a specific sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            List of FormulaInfo objects
        """
        sheet = self.workbook[sheet_name]
        formulas = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula cell
                    formula_info = self._analyze_formula(
                        sheet_name=sheet_name,
                        cell_address=cell.coordinate,
                        formula=cell.value
                    )
                    if formula_info:
                        formulas.append(formula_info)
        
        return formulas
    
    def _analyze_formula(self, sheet_name: str, cell_address: str, formula: str) -> Optional[FormulaInfo]:
        """
        Analyze a formula to extract metadata.
        
        Args:
            sheet_name: Name of the sheet
            cell_address: Cell address (e.g., 'A1')
            formula: The formula string
            
        Returns:
            FormulaInfo object or None if analysis fails
        """
        if not formula:
            return None
        
        # Determine formula type
        formula_type = 'standard'
        if formula.startswith('{') and formula.endswith('}'):
            formula_type = 'array'
        
        # Extract dependencies (referenced cells and ranges)
        dependencies = self._extract_dependencies(formula)
        
        # Extract Excel functions used
        functions_used = self._extract_functions(formula)
        
        return FormulaInfo(
            sheet_name=sheet_name,
            cell_address=cell_address,
            formula=formula,
            formula_type=formula_type,
            dependencies=dependencies,
            contains_functions=functions_used
        )
    
    def _extract_dependencies(self, formula: str) -> List[str]:
        """
        Extract cell references and ranges from a formula.
        
        Args:
            formula: The formula string
            
        Returns:
            List of cell references and ranges
        """
        dependencies = []
        
        # Pattern for cell references (e.g., A1, $A$1, Sheet1!A1)
        cell_pattern = r"(?:['\"]?[\w\s]+['\"]?!)?\$?[A-Z]+\$?\d+"
        
        # Pattern for range references (e.g., A1:B10)
        range_pattern = r"(?:['\"]?[\w\s]+['\"]?!)?\$?[A-Z]+\$?\d+:\$?[A-Z]+\$?\d+"
        
        # Extract ranges first (they contain cell patterns)
        ranges = re.findall(range_pattern, formula, re.IGNORECASE)
        dependencies.extend(ranges)
        
        # Extract individual cells (excluding those in ranges)
        formula_without_ranges = formula
        for range_ref in ranges:
            formula_without_ranges = formula_without_ranges.replace(range_ref, '')
        
        cells = re.findall(cell_pattern, formula_without_ranges, re.IGNORECASE)
        dependencies.extend(cells)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_deps = []
        for dep in dependencies:
            if dep not in seen:
                seen.add(dep)
                unique_deps.append(dep)
        
        return unique_deps
    
    def _extract_functions(self, formula: str) -> List[str]:
        """
        Extract Excel function names from a formula.
        
        Args:
            formula: The formula string
            
        Returns:
            List of function names used
        """
        functions_found = []
        
        # Pattern for function names: uppercase letters followed by (
        function_pattern = r'\b([A-Z][A-Z0-9_]*)\s*\('
        
        matches = re.findall(function_pattern, formula)
        
        for func_name in matches:
            if func_name in self.EXCEL_FUNCTIONS:
                functions_found.append(func_name)
        
        return list(set(functions_found))  # Remove duplicates
    
    def get_formulas_by_sheet(self, formulas: List[FormulaInfo]) -> Dict[str, List[FormulaInfo]]:
        """
        Organize formulas by sheet name.
        
        Args:
            formulas: List of FormulaInfo objects
            
        Returns:
            Dictionary mapping sheet names to their formulas
        """
        by_sheet = {}
        for formula_info in formulas:
            if formula_info.sheet_name not in by_sheet:
                by_sheet[formula_info.sheet_name] = []
            by_sheet[formula_info.sheet_name].append(formula_info)
        return by_sheet
    
    def get_formulas_by_function(self, formulas: List[FormulaInfo]) -> Dict[str, List[FormulaInfo]]:
        """
        Organize formulas by the Excel functions they use.
        
        Args:
            formulas: List of FormulaInfo objects
            
        Returns:
            Dictionary mapping function names to formulas that use them
        """
        by_function = {}
        for formula_info in formulas:
            for func_name in formula_info.contains_functions:
                if func_name not in by_function:
                    by_function[func_name] = []
                by_function[func_name].append(formula_info)
        return by_function
    
    def get_formula_statistics(self, formulas: List[FormulaInfo]) -> Dict:
        """
        Generate statistics about the extracted formulas.
        
        Args:
            formulas: List of FormulaInfo objects
            
        Returns:
            Dictionary containing statistics
        """
        all_functions = []
        for formula_info in formulas:
            all_functions.extend(formula_info.contains_functions)
        
        function_counts = {}
        for func in all_functions:
            function_counts[func] = function_counts.get(func, 0) + 1
        
        return {
            'total_formulas': len(formulas),
            'sheets_with_formulas': len({f.sheet_name for f in formulas}),
            'unique_functions_used': len(set(all_functions)),
            'function_usage': function_counts,
            'array_formulas': sum(1 for f in formulas if f.formula_type == 'array'),
            'most_common_functions': sorted(
                function_counts.items(), 
                key=lambda x: x[1], 
                reverse=True
            )[:10]
        }
