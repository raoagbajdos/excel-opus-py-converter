"""
Data Exporter Module
Exports Excel data to pandas DataFrames and generates Python code
"""
import logging
import os
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class SheetData:
    """Information about exported sheet data."""
    sheet_name: str
    dataframe: pd.DataFrame
    data_range: str  # e.g., "A1:F100"
    has_header: bool
    dtypes: Dict[str, str]  # Column name to dtype mapping
    
    
@dataclass
class ExportResult:
    """Result of data export operation."""
    sheet_data: List[SheetData]
    python_code: str  # Generated Python code to recreate DataFrames
    metadata: Dict[str, Any]


class DataExporter:
    """Export Excel data to pandas DataFrames and generate Python code."""
    
    def __init__(self, filepath: str):
        """
        Initialize the data exporter.
        
        Args:
            filepath: Path to the Excel file
        """
        self.filepath = filepath
        self.workbook: Optional[openpyxl.Workbook] = None
        
    def export_all_sheets(self, 
                          include_empty: bool = False,
                          infer_header: bool = True,
                          max_rows: Optional[int] = None) -> ExportResult:
        """
        Export all sheets to DataFrames.
        
        Args:
            include_empty: Include sheets with no data
            infer_header: Try to detect header rows
            max_rows: Maximum rows to export per sheet (None = all)
            
        Returns:
            ExportResult containing all sheet data and Python code
        """
        ext = os.path.splitext(self.filepath)[1].lower()

        # For genuine .xls (OLE/BIFF) files, use pandas with xlrd engine
        if ext == '.xls':
            return self._export_with_pandas_xlrd(
                include_empty=include_empty,
                infer_header=infer_header,
                max_rows=max_rows,
            )

        self.workbook = openpyxl.load_workbook(self.filepath, data_only=True)
        sheet_data_list = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet_data = self._export_sheet(
                sheet_name=sheet_name,
                infer_header=infer_header,
                max_rows=max_rows
            )
            
            if sheet_data and (include_empty or not sheet_data.dataframe.empty):
                sheet_data_list.append(sheet_data)
        
        self.workbook.close()
        
        # Generate Python code
        python_code = self._generate_python_code(sheet_data_list)
        
        # Generate metadata
        metadata = self._generate_metadata(sheet_data_list)
        
        return ExportResult(
            sheet_data=sheet_data_list,
            python_code=python_code,
            metadata=metadata
        )

    def _export_with_pandas_xlrd(self, *,
                                  include_empty: bool,
                                  infer_header: bool,
                                  max_rows: Optional[int]) -> ExportResult:
        """Read a genuine .xls file via pandas (xlrd engine)."""
        try:
            all_sheets = pd.read_excel(
                self.filepath, sheet_name=None, header=0 if infer_header else None,
                nrows=max_rows, engine='xlrd',
            )
        except Exception as exc:
            logger.warning("pandas/xlrd could not read %s: %s", self.filepath, exc)
            return ExportResult(
                sheet_data=[], python_code='# Could not read .xls file', metadata=self._generate_metadata([]),
            )

        sheet_data_list: List[SheetData] = []
        for sheet_name, df in all_sheets.items():
            if not include_empty and df.empty:
                continue
            dtypes = {col: str(df[col].dtype) for col in df.columns}
            nrows, ncols = df.shape
            data_range = f"A1:{get_column_letter(max(ncols, 1))}{nrows + 1}"
            sheet_data_list.append(SheetData(
                sheet_name=str(sheet_name),
                dataframe=df,
                data_range=data_range,
                has_header=infer_header,
                dtypes=dtypes,
            ))

        return ExportResult(
            sheet_data=sheet_data_list,
            python_code=self._generate_python_code(sheet_data_list),
            metadata=self._generate_metadata(sheet_data_list),
        )
    
    def _export_sheet(self,
                      sheet_name: str,
                      infer_header: bool = True,
                      max_rows: Optional[int] = None) -> Optional[SheetData]:
        """
        Export a single sheet to DataFrame.
        
        Args:
            sheet_name: Name of the sheet
            infer_header: Try to detect header rows
            max_rows: Maximum rows to export
            
        Returns:
            SheetData object or None if sheet is empty
        """
        sheet = self.workbook[sheet_name]
        
        # Find used range
        data_range_info = self._get_used_range(sheet)
        if not data_range_info:
            return None
        
        min_row, max_row, min_col, max_col = data_range_info
        
        # Apply max_rows limit if specified
        if max_rows:
            max_row = min(max_row, min_row + max_rows - 1)
        
        # Extract data
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True
        ), start=min_row):
            data.append(list(row))
        
        if not data:
            return None
        
        # Check if first row looks like a header
        has_header = False
        if infer_header and len(data) > 1:
            has_header = self._looks_like_header(data[0], data[1:])
        
        # Create DataFrame
        if has_header:
            df = pd.DataFrame(data[1:], columns=data[0])
        else:
            df = pd.DataFrame(data)
        
        # Clean up column names if they exist
        if has_header:
            df.columns = [self._clean_column_name(str(col)) for col in df.columns]
        
        # Get dtype information
        dtypes = {col: str(df[col].dtype) for col in df.columns}
        
        # Format data range string
        data_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        
        return SheetData(
            sheet_name=sheet_name,
            dataframe=df,
            data_range=data_range,
            has_header=has_header,
            dtypes=dtypes
        )
    
    def _get_used_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """
        Find the actual used range in a sheet (ignoring empty rows/columns).
        
        Args:
            sheet: openpyxl worksheet
            
        Returns:
            Tuple of (min_row, max_row, min_col, max_col) or None if empty
        """
        if sheet.max_row is None or sheet.max_column is None:
            return None
        
        min_row = self._find_first_nonempty_row(sheet)
        if min_row is None:
            return None
        
        max_row = self._find_last_nonempty_row(sheet)
        min_col = self._find_boundary_col(sheet, min_row, max_row, reverse=False)
        max_col = self._find_boundary_col(sheet, min_row, max_row, reverse=True)
        
        if all(v is not None for v in [min_row, max_row, min_col, max_col]):
            return (min_row, max_row, min_col, max_col)
        return None

    @staticmethod
    def _find_first_nonempty_row(sheet) -> Optional[int]:
        """Return the first row index containing a non-empty cell."""
        for row_idx in range(1, sheet.max_row + 1):
            if any(cell.value is not None for cell in sheet[row_idx]):
                return row_idx
        return None

    @staticmethod
    def _find_last_nonempty_row(sheet) -> Optional[int]:
        """Return the last row index containing a non-empty cell."""
        for row_idx in range(sheet.max_row, 0, -1):
            if any(cell.value is not None for cell in sheet[row_idx]):
                return row_idx
        return None

    @staticmethod
    def _find_boundary_col(sheet, min_row: int, max_row: int,
                           *, reverse: bool = False) -> Optional[int]:
        """Return the first (or last when *reverse*) non-empty column index."""
        col_range = (
            range(sheet.max_column, 0, -1) if reverse
            else range(1, sheet.max_column + 1)
        )
        for col_idx in col_range:
            col_values = [sheet.cell(row=r, column=col_idx).value
                         for r in range(min_row, max_row + 1)]
            if any(val is not None for val in col_values):
                return col_idx
        return None
    
    def _looks_like_header(self, first_row: List[Any], data_rows: List[List[Any]]) -> bool:
        """
        Determine if the first row looks like a header.
        
        Args:
            first_row: First row values
            data_rows: Remaining data rows
            
        Returns:
            True if first row appears to be a header
        """
        if not first_row or not data_rows:
            return False
        
        # Check 1: All values in first row are non-empty
        if not all(val is not None and str(val).strip() for val in first_row):
            return False
        
        # Check 2: First row contains mostly strings
        string_count = sum(1 for val in first_row if isinstance(val, str))
        if string_count / len(first_row) < 0.5:
            return False
        
        # Check 3: First row types differ from data row types
        if len(data_rows) > 0:
            first_data_row = data_rows[0]
            type_matches = sum(
                1 for h_val, d_val in zip(first_row, first_data_row)
                if type(h_val) == type(d_val) and h_val is not None and d_val is not None
            )
            if type_matches > len(first_row) * 0.7:
                return False
        
        return True
    
    def _clean_column_name(self, name: str) -> str:
        """
        Clean up column names to be valid Python identifiers.
        
        Args:
            name: Original column name
            
        Returns:
            Cleaned column name
        """
        if not name or name == 'None':
            return 'Unnamed'
        
        # Replace spaces and special characters with underscores
        clean = str(name).strip()
        clean = clean.replace(' ', '_').replace('-', '_')
        
        # Remove non-alphanumeric characters (except underscores)
        clean = ''.join(c if c.isalnum() or c == '_' else '_' for c in clean)
        
        # Ensure it starts with a letter or underscore
        if clean and not (clean[0].isalpha() or clean[0] == '_'):
            clean = 'col_' + clean
        
        return clean or 'Unnamed'
    
    def _generate_python_code(self, sheet_data_list: List[SheetData]) -> str:
        """
        Generate Python code to recreate the exported DataFrames.
        
        Args:
            sheet_data_list: List of SheetData objects
            
        Returns:
            Python code as string
        """
        code_lines = [
            '"""',
            'Generated Python code to recreate Excel data as pandas DataFrames',
            '"""',
            'import pandas as pd',
            'import numpy as np',
            'from pathlib import Path',
            '',
            ''
        ]
        
        # Generate code for each sheet
        for sheet_data in sheet_data_list:
            # Create safe variable name from sheet name
            var_name = self._clean_column_name(sheet_data.sheet_name).lower()
            
            code_lines.append(f'# Data from sheet: {sheet_data.sheet_name}')
            code_lines.append(f'# Original range: {sheet_data.data_range}')
            
            # Option 1: Read from Excel file
            code_lines.append(f'{var_name}_df = pd.read_excel(')
            code_lines.append('    "your_file.xlsx",')
            code_lines.append(f'    sheet_name="{sheet_data.sheet_name}",')
            if sheet_data.has_header:
                code_lines.append('    header=0')
            else:
                code_lines.append('    header=None')
            code_lines.append(')')
            code_lines.append('')
            
            # Option 2: Create from dictionary (for small datasets)
            if len(sheet_data.dataframe) <= 10:
                code_lines.append(f'# Alternative: Create {var_name}_df from data:')
                data_dict = sheet_data.dataframe.to_dict('list')
                code_lines.append(f'{var_name}_data = {repr(data_dict)}')
                code_lines.append(f'{var_name}_df_alt = pd.DataFrame({var_name}_data)')
                code_lines.append('')
            
            # Show DataFrame info
            code_lines.append(f'print(f"Shape of {sheet_data.sheet_name}: {{{var_name}_df.shape}}")')
            code_lines.append(f'print({var_name}_df.head())')
            code_lines.append('')
            code_lines.append('')
        
        # Add summary
        code_lines.append('# Summary of all DataFrames')
        code_lines.append('dataframes = {')
        for sheet_data in sheet_data_list:
            var_name = self._clean_column_name(sheet_data.sheet_name).lower()
            code_lines.append(f'    "{sheet_data.sheet_name}": {var_name}_df,')
        code_lines.append('}')
        code_lines.append('')
        code_lines.append('for name, df in dataframes.items():')
        code_lines.append('    print(f"\\n{name}: {df.shape[0]} rows, {df.shape[1]} columns")')
        
        return '\n'.join(code_lines)
    
    def _generate_metadata(self, sheet_data_list: List[SheetData]) -> Dict[str, Any]:
        """
        Generate metadata about the exported data.
        
        Args:
            sheet_data_list: List of SheetData objects
            
        Returns:
            Metadata dictionary
        """
        total_rows = sum(len(sd.dataframe) for sd in sheet_data_list)
        total_cols = sum(len(sd.dataframe.columns) for sd in sheet_data_list)
        
        return {
            'total_sheets': len(sheet_data_list),
            'total_rows': total_rows,
            'total_columns': total_cols,
            'sheets': [
                {
                    'name': sd.sheet_name,
                    'rows': len(sd.dataframe),
                    'columns': len(sd.dataframe.columns),
                    'data_range': sd.data_range,
                    'has_header': sd.has_header,
                    'column_names': list(sd.dataframe.columns) if sd.has_header else None,
                    'dtypes': sd.dtypes
                }
                for sd in sheet_data_list
            ]
        }
    
    def export_to_csv(self, sheet_data: SheetData, output_path: str):
        """
        Export a sheet's data to CSV.
        
        Args:
            sheet_data: SheetData object
            output_path: Path for output CSV file
        """
        sheet_data.dataframe.to_csv(output_path, index=False)
    
    def export_to_json(self, sheet_data: SheetData, output_path: str):
        """
        Export a sheet's data to JSON.
        
        Args:
            sheet_data: SheetData object
            output_path: Path for output JSON file
        """
        sheet_data.dataframe.to_json(output_path, orient='records', indent=2)
