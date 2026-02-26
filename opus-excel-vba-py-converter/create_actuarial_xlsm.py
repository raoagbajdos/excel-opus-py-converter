#!/usr/bin/env python3
"""
Generate an actuarial calculations .xlsm workbook with EMBEDDED VBA macros.

This script creates a proper vbaProject.bin using the MS-OVBA specification
and injects it into an .xlsm file so that oletools/olevba can extract the macros.

Usage:
    python create_actuarial_xlsm.py [output_path]
"""
from __future__ import annotations

import os
import struct
import sys
import zlib
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# ============================================================================
# VBA PROJECT BINARY BUILDER  (MS-OVBA / MS-CFB specification)
# ============================================================================

def _compress_vba(source: bytes) -> bytes:
    """Compress VBA source using the MS-OVBA RLE compression algorithm."""
    compressed = bytearray(b"\x01")  # signature byte
    src_pos = 0
    src_len = len(source)

    while src_pos < src_len:
        # Start a new chunk (max 4096 bytes decompressed)
        chunk_start = src_pos
        chunk_data = bytearray()
        chunk_data.append(0)  # placeholder for header
        chunk_data.append(0)

        flags_pos = len(chunk_data)
        flags = 0
        flag_bit = 0

        # We need a flags byte every 8 tokens
        chunk_data.append(0)  # placeholder for first flags byte

        while src_pos < src_len and (src_pos - chunk_start) < 4096:
            if flag_bit == 8:
                # Write accumulated flags and start new flags byte
                chunk_data[flags_pos] = flags
                flags = 0
                flag_bit = 0
                flags_pos = len(chunk_data)
                chunk_data.append(0)

            # Try to find a match in the decompressed buffer
            best_len = 0
            best_off = 0
            decompressed_cur = src_pos - chunk_start
            # bit_count determines offset/length split
            if decompressed_cur <= 0x10:
                max_off_bits = 4
            elif decompressed_cur <= 0x20:
                max_off_bits = 5
            elif decompressed_cur <= 0x40:
                max_off_bits = 6
            elif decompressed_cur <= 0x80:
                max_off_bits = 7
            elif decompressed_cur <= 0x100:
                max_off_bits = 8
            elif decompressed_cur <= 0x200:
                max_off_bits = 9
            elif decompressed_cur <= 0x400:
                max_off_bits = 10
            elif decompressed_cur <= 0x800:
                max_off_bits = 11
            else:
                max_off_bits = 12

            len_bits = 16 - max_off_bits
            max_match = (1 << len_bits) - 1 + 3
            max_off = (1 << max_off_bits)

            if decompressed_cur > 0:
                search_start = max(0, decompressed_cur - max_off)
                for off in range(search_start, decompressed_cur):
                    match_len = 0
                    while (match_len < max_match
                           and (decompressed_cur + match_len) < 4096
                           and (src_pos + match_len) < src_len
                           and source[chunk_start + off + match_len] == source[src_pos + match_len]):
                        match_len += 1
                    if match_len > best_len:
                        best_len = match_len
                        best_off = decompressed_cur - off

            if best_len >= 3:
                # Emit copy token
                offset_encoded = best_off - 1
                length_encoded = best_len - 3
                token = (offset_encoded << len_bits) | length_encoded
                chunk_data.append(token & 0xFF)
                chunk_data.append((token >> 8) & 0xFF)
                flags |= (1 << flag_bit)
                src_pos += best_len
            else:
                # Emit literal byte
                chunk_data.append(source[src_pos])
                src_pos += 1

            flag_bit += 1

        # Write last flags byte
        chunk_data[flags_pos] = flags

        # Write chunk header: bit 15 = 1 (compressed), bits 0-11 = size - 3
        chunk_size = len(chunk_data) - 2  # exclude the 2-byte header
        chunk_header = 0xB000 | (chunk_size - 1)
        chunk_data[0] = chunk_header & 0xFF
        chunk_data[1] = (chunk_header >> 8) & 0xFF

        compressed.extend(chunk_data)

    return bytes(compressed)


def _build_dir_stream(project_name: str, modules: dict[str, str]) -> bytes:
    """Build the compressed 'dir' stream for the VBA project."""
    buf = BytesIO()

    def _write_record(record_id: int, data: bytes) -> None:
        buf.write(struct.pack("<HI", record_id, len(data)))
        buf.write(data)

    # PROJECTSYSKIND
    _write_record(0x0001, struct.pack("<I", 0x00000001))  # Win32
    # PROJECTLCID
    _write_record(0x0002, struct.pack("<I", 0x0409))
    # PROJECTLCIDINVOKE
    _write_record(0x0014, struct.pack("<I", 0x0409))
    # PROJECTCODEPAGE
    _write_record(0x0003, struct.pack("<H", 1252))
    # PROJECTNAME
    _write_record(0x0004, project_name.encode("ascii"))
    # PROJECTDOCSTRING
    _write_record(0x0005, b"Actuarial VBA Project")
    _write_record(0x0040, b"")  # unicode docstring
    # PROJECTHELPFILE
    _write_record(0x0006, b"")
    _write_record(0x003D, b"")
    # PROJECTHELPCONTEXT
    _write_record(0x0007, struct.pack("<I", 0))
    # PROJECTLIBFLAGS
    _write_record(0x0008, struct.pack("<I", 0))
    # PROJECTVERSION
    buf.write(struct.pack("<HI", 0x0009, 4))
    buf.write(struct.pack("<IH", 1, 1))
    # PROJECTCONSTANTS
    _write_record(0x000C, b"")
    _write_record(0x003C, b"")

    # === REFERENCES ===
    # Reference to stdole
    ref_name = b"stdole"
    _write_record(0x0016, ref_name)  # REFERENCENAME
    # REFERENCEREGISTERED
    libid = b"*\\G{00020430-0000-0000-C000-000000000046}#2.0#0#C:\\Windows\\System32\\stdole2.tlb#OLE Automation"
    buf.write(struct.pack("<HI", 0x000D, len(libid) + 4))
    buf.write(struct.pack("<I", len(libid)))
    buf.write(libid)

    # Reference to Office
    ref_name2 = b"Office"
    _write_record(0x0016, ref_name2)
    libid2 = b"*\\G{2DF8D04C-5BFA-101B-BDE5-00AA0044DE52}#2.0#0#C:\\Program Files\\Common Files\\Microsoft Shared\\OFFICE16\\MSO.DLL#Microsoft Office 16.0 Object Library"
    buf.write(struct.pack("<HI", 0x000D, len(libid2) + 4))
    buf.write(struct.pack("<I", len(libid2)))
    buf.write(libid2)

    # === MODULES ===
    # Project module count
    _write_record(0x000F, struct.pack("<H", len(modules) + 1))  # +1 for ThisWorkbook
    # PROJECTCOOKIE
    _write_record(0x0013, struct.pack("<H", 0xFFFF))

    # Module: ThisWorkbook (document module)
    tw_code = "Attribute VB_Name = \"ThisWorkbook\"\nAttribute VB_Base = \"0{00020819-0000-0000-C000-000000000046}\"\nAttribute VB_GlobalNameSpace = False\nAttribute VB_Creatable = False\nAttribute VB_PredeclaredId = True\nAttribute VB_Exposed = True\nAttribute VB_TemplateDerived = False\nAttribute VB_Customizable = True\n"
    _write_module(buf, "ThisWorkbook", tw_code, is_document=True)

    # Standard modules
    for mod_name, mod_code in modules.items():
        full_code = f'Attribute VB_Name = "{mod_name}"\n{mod_code}'
        _write_module(buf, mod_name, full_code, is_document=False)

    # End of modules
    buf.write(struct.pack("<HI", 0x0010, 0))

    raw = buf.getvalue()
    return _compress_vba(raw)


def _write_module(buf: BytesIO, name: str, code: str, *, is_document: bool) -> None:
    """Write a single module record into the dir stream."""
    name_bytes = name.encode("ascii")

    def _write_record(record_id: int, data: bytes) -> None:
        buf.write(struct.pack("<HI", record_id, len(data)))
        buf.write(data)

    _write_record(0x0019, name_bytes)                       # MODULENAME
    _write_record(0x0047, name.encode("utf-16-le"))         # MODULENAMEUNICODE
    _write_record(0x001C, name_bytes)                       # MODULESTREAMNAME
    _write_record(0x0032, name.encode("utf-16-le"))         # MODULESTREAMNAME unicode
    _write_record(0x001A, b"")                              # MODULEDOCSTRING
    _write_record(0x0048, b"")                              # MODULEDOCSTRING unicode
    _write_record(0x0031, struct.pack("<I", 0))             # MODULEOFFSET
    _write_record(0x001E, struct.pack("<I", 0))             # MODULEHELPCONTEXT
    _write_record(0x002C, struct.pack("<I", 0))             # MODULECOOKIE
    if is_document:
        _write_record(0x0022, struct.pack("<I", 0))         # MODULETYPEDOCUMENT
    else:
        _write_record(0x0021, struct.pack("<I", 0))         # MODULETYPEPROCEDURAL
    # MODULEREADONLY is omitted
    # MODULEPRIVATE is omitted
    buf.write(struct.pack("<HI", 0x002B, 0))                # MODULE terminator


def build_vba_project_bin(project_name: str, modules: dict[str, str]) -> bytes:
    """
    Build a complete vbaProject.bin as an OLE Compound File.

    Uses the olefile library to write the OLE structure, then populates
    the required VBA streams.
    """
    import tempfile

    # Build compressed module streams
    module_streams: dict[str, bytes] = {}

    # ThisWorkbook (empty document module)
    tw_src = 'Attribute VB_Name = "ThisWorkbook"\r\nAttribute VB_Base = "0{00020819-0000-0000-C000-000000000046}"\r\nAttribute VB_GlobalNameSpace = False\r\nAttribute VB_Creatable = False\r\nAttribute VB_PredeclaredId = True\r\nAttribute VB_Exposed = True\r\nAttribute VB_TemplateDerived = False\r\nAttribute VB_Customizable = True\r\n'
    module_streams["ThisWorkbook"] = _compress_vba(tw_src.encode("ascii"))

    for mod_name, code in modules.items():
        full_code = f'Attribute VB_Name = "{mod_name}"\r\n{code}'
        module_streams[mod_name] = _compress_vba(full_code.encode("ascii"))

    # Build dir stream
    dir_stream = _build_dir_stream(project_name, modules)

    # _VBA_PROJECT stream (minimal header)
    vba_project_stream = struct.pack("<HH", 0x61CC, 0x0000) + b"\x00" * 3

    # PROJECT stream (text metadata)
    project_lines = [
        f'ID="{{00000000-0000-0000-0000-000000000001}}"',
        f'Document=ThisWorkbook/&H00000000',
    ]
    for mod_name in modules:
        project_lines.append(f'Module={mod_name}')
    project_lines.extend([
        f'Name="{project_name}"',
        f'HelpContextID="0"',
        f'VersionCompatible32="393222000"',
        f'CMG="0000"',
        f'DPB="0000"',
        f'GC="0000"',
        '',
        '[Host Extender Info]',
        '&H00000001={3832D640-CF90-11CF-8E43-00A0C911005A};VBE;&H00000000',
        '',
        '[Workspace]',
        f'ThisWorkbook=0, 0, 0, 0, C',
    ])
    for mod_name in modules:
        project_lines.append(f'{mod_name}=0, 0, 0, 0, C')
    project_stream = "\r\n".join(project_lines).encode("ascii")

    # PROJECTwm stream (unicode name map)
    projectwm = bytearray()
    all_names = ["ThisWorkbook"] + list(modules.keys())
    for name in all_names:
        projectwm.extend(name.encode("ascii"))
        projectwm.append(0)
        projectwm.extend(name.encode("utf-16-le"))
        projectwm.extend(b"\x00\x00")
    projectwm.append(0)  # terminal null

    # Now build the OLE compound file
    # olefile doesn't support writing, so we build the CFB manually
    # using a simplified approach based on the MS-CFB specification
    return _build_cfb(
        dir_stream=dir_stream,
        vba_project_stream=vba_project_stream,
        module_streams=module_streams,
        project_stream=project_stream,
        projectwm_stream=bytes(projectwm),
    )


def _build_cfb(
    dir_stream: bytes,
    vba_project_stream: bytes,
    module_streams: dict[str, bytes],
    project_stream: bytes,
    projectwm_stream: bytes,
) -> bytes:
    """Build a minimal Compound File Binary (MS-CFB) containing VBA streams."""
    SECTOR_SIZE = 512
    MINI_SECTOR_SIZE = 64
    MINI_STREAM_CUTOFF = 0x1000
    ENDOFCHAIN = 0xFFFFFFFE
    FREESECT = 0xFFFFFFFF
    NOSTREAM = 0xFFFFFFFF

    # Collect all streams that go into the mini-stream
    # (all our streams are small enough for mini-stream)
    entries: list[dict] = []
    # Index 0: Root Entry (always first)
    # Index 1: VBA storage
    # Index 2: _VBA_PROJECT (stream under VBA/)
    # Index 3: dir (stream under VBA/)
    # Index 4+ : module streams under VBA/
    # Then: PROJECT, PROJECTwm at root level

    vba_child_entries: list[dict] = []
    vba_child_entries.append({
        "name": "_VBA_PROJECT",
        "type": 2,  # stream
        "data": vba_project_stream,
    })
    vba_child_entries.append({
        "name": "dir",
        "type": 2,
        "data": dir_stream,
    })
    for mod_name, mod_data in module_streams.items():
        vba_child_entries.append({
            "name": mod_name,
            "type": 2,
            "data": mod_data,
        })

    root_child_entries: list[dict] = []
    root_child_entries.append({
        "name": "VBA",
        "type": 1,  # storage
        "children": vba_child_entries,
    })
    root_child_entries.append({
        "name": "PROJECT",
        "type": 2,
        "data": project_stream,
    })
    root_child_entries.append({
        "name": "PROJECTwm",
        "type": 2,
        "data": projectwm_stream,
    })

    # Flatten into directory entries list
    dir_entries: list[dict] = []

    def add_entry(name: str, entry_type: int, data: bytes = b"",
                  child_id: int = NOSTREAM, left: int = NOSTREAM,
                  right: int = NOSTREAM, color: int = 1) -> int:
        idx = len(dir_entries)
        dir_entries.append({
            "name": name,
            "type": entry_type,  # 5=root, 1=storage, 2=stream
            "data": data,
            "child_id": child_id,
            "left_sibling": left,
            "right_sibling": right,
            "color": color,  # 0=red, 1=black
        })
        return idx

    # Build flat directory using red-black tree (simplified: linear chain)
    # Root entry
    root_idx = add_entry("Root Entry", 5)

    # VBA storage
    vba_idx = add_entry("VBA", 1)

    # VBA children
    vba_first_child = len(dir_entries)
    for i, child in enumerate(vba_child_entries):
        add_entry(child["name"], 2, data=child["data"])

    # Arrange VBA children as balanced tree
    _arrange_siblings(dir_entries, vba_first_child, len(vba_child_entries))
    dir_entries[vba_idx]["child_id"] = vba_first_child + len(vba_child_entries) // 2

    # PROJECT stream
    proj_idx = add_entry("PROJECT", 2, data=project_stream)
    # PROJECTwm stream
    projwm_idx = add_entry("PROJECTwm", 2, data=projectwm_stream)

    # Arrange root children (VBA, PROJECT, PROJECTwm) as tree
    # Sort by name for red-black tree: PROJECT < PROJECTwm < VBA
    # Middle = PROJECTwm, left = PROJECT, right = VBA
    dir_entries[proj_idx]["left_sibling"] = NOSTREAM
    dir_entries[proj_idx]["right_sibling"] = NOSTREAM
    dir_entries[proj_idx]["color"] = 0  # red

    dir_entries[projwm_idx]["left_sibling"] = proj_idx
    dir_entries[projwm_idx]["right_sibling"] = vba_idx
    dir_entries[projwm_idx]["color"] = 1  # black

    dir_entries[vba_idx]["left_sibling"] = NOSTREAM
    dir_entries[vba_idx]["right_sibling"] = NOSTREAM
    dir_entries[vba_idx]["color"] = 0  # red

    dir_entries[root_idx]["child_id"] = projwm_idx

    # Build mini-stream from all stream data
    mini_stream = bytearray()
    mini_fat: list[int] = []

    for entry in dir_entries:
        if entry["type"] == 2 and entry.get("data"):
            data = entry["data"]
            start_sector = len(mini_fat)
            entry["mini_start"] = start_sector
            entry["size"] = len(data)

            # Pad to mini-sector boundary
            padded = data + b"\x00" * (MINI_SECTOR_SIZE - len(data) % MINI_SECTOR_SIZE) if len(data) % MINI_SECTOR_SIZE else data
            n_sectors = len(padded) // MINI_SECTOR_SIZE

            for j in range(n_sectors):
                mini_stream.extend(padded[j * MINI_SECTOR_SIZE:(j + 1) * MINI_SECTOR_SIZE])
                if j < n_sectors - 1:
                    mini_fat.append(start_sector + j + 1)
                else:
                    mini_fat.append(ENDOFCHAIN)
        else:
            entry["mini_start"] = ENDOFCHAIN if entry["type"] != 5 else 0
            entry["size"] = 0

    # Root entry size = mini-stream total size
    dir_entries[root_idx]["size"] = len(mini_stream)

    # Build the regular sectors:
    # 1. Mini-stream sectors
    # 2. Mini-FAT sectors
    # 3. Directory sectors
    # 4. FAT sectors
    # 5. DIFAT sectors (if needed)

    sectors: list[bytes] = []

    # Mini-stream sectors
    mini_stream_start = len(sectors)
    padded_mini = bytes(mini_stream) + b"\x00" * (SECTOR_SIZE - len(mini_stream) % SECTOR_SIZE) if len(mini_stream) % SECTOR_SIZE else bytes(mini_stream)
    if not padded_mini:
        padded_mini = b"\x00" * SECTOR_SIZE
    n_mini_sectors = max(1, len(padded_mini) // SECTOR_SIZE)
    for i in range(n_mini_sectors):
        sectors.append(padded_mini[i * SECTOR_SIZE:(i + 1) * SECTOR_SIZE])

    dir_entries[root_idx]["mini_start"] = mini_stream_start

    # Mini-FAT sector
    mini_fat_start = len(sectors)
    mini_fat_data = b"".join(struct.pack("<I", x) for x in mini_fat)
    mini_fat_data += b"\xff\xff\xff\xff" * (SECTOR_SIZE // 4 - len(mini_fat))
    mini_fat_data = mini_fat_data[:SECTOR_SIZE]
    sectors.append(mini_fat_data)

    # Directory sector(s)
    dir_start = len(sectors)
    dir_data = bytearray()
    for entry in dir_entries:
        dir_data.extend(_build_dir_entry(entry, NOSTREAM))
    # Pad to sector boundary
    while len(dir_data) % SECTOR_SIZE:
        dir_data.extend(b"\x00" * 128)  # empty dir entries
    n_dir_sectors = len(dir_data) // SECTOR_SIZE
    for i in range(n_dir_sectors):
        sectors.append(bytes(dir_data[i * SECTOR_SIZE:(i + 1) * SECTOR_SIZE]))

    # Now build FAT
    fat_start = len(sectors)
    n_fat_sectors = 1  # usually enough for small files
    # Reserve space for FAT sector(s)
    total_sectors = len(sectors) + n_fat_sectors

    fat: list[int] = []
    # Mini-stream chain
    for i in range(n_mini_sectors):
        if i < n_mini_sectors - 1:
            fat.append(mini_stream_start + i + 1)
        else:
            fat.append(ENDOFCHAIN)

    # Mini-FAT: single sector
    fat.append(ENDOFCHAIN)

    # Directory chain
    for i in range(n_dir_sectors):
        if i < n_dir_sectors - 1:
            fat.append(dir_start + i + 1)
        else:
            fat.append(ENDOFCHAIN)

    # FAT sector itself
    fat.append(0xFFFFFFFD)  # FATSECT marker

    # Pad FAT to fill sector
    while len(fat) < SECTOR_SIZE // 4:
        fat.append(FREESECT)

    fat_data = b"".join(struct.pack("<I", x) for x in fat[:SECTOR_SIZE // 4])
    sectors.append(fat_data)

    # Build header
    header = _build_cfb_header(
        total_sectors=len(sectors),
        fat_sectors=[fat_start],
        dir_start=dir_start,
        mini_fat_start=mini_fat_start,
        n_mini_fat_sectors=1,
    )

    # Assemble file
    result = bytearray(header)
    for sector in sectors:
        assert len(sector) == SECTOR_SIZE, f"Sector size mismatch: {len(sector)}"
        result.extend(sector)

    return bytes(result)


def _arrange_siblings(entries: list[dict], start: int, count: int) -> None:
    """Arrange entries as a minimal red-black tree (simplified)."""
    NOSTREAM = 0xFFFFFFFF
    if count <= 0:
        return

    mid = start + count // 2
    for i in range(start, start + count):
        entries[i]["left_sibling"] = NOSTREAM
        entries[i]["right_sibling"] = NOSTREAM
        entries[i]["color"] = 1  # black

    if count == 1:
        return
    elif count == 2:
        entries[start]["right_sibling"] = NOSTREAM
        entries[start]["left_sibling"] = NOSTREAM
        entries[start + 1]["left_sibling"] = start
        entries[start + 1]["right_sibling"] = NOSTREAM
        entries[start]["color"] = 0  # red
    elif count >= 3:
        entries[mid]["left_sibling"] = start + (mid - start) // 2 if mid > start else NOSTREAM
        entries[mid]["right_sibling"] = mid + 1 + (start + count - mid - 1) // 2 if mid + 1 < start + count else NOSTREAM
        # Recursively arrange left and right subtrees (simplified: just set leaves)
        for i in range(start, mid):
            entries[i]["color"] = 0  # red (leaves)
        for i in range(mid + 1, start + count):
            entries[i]["color"] = 0  # red (leaves)


def _build_dir_entry(entry: dict, nostream: int) -> bytes:
    """Build a 128-byte directory entry for the CFB."""
    name = entry["name"]
    name_utf16 = name.encode("utf-16-le") + b"\x00\x00"
    name_size = len(name_utf16)

    buf = bytearray(128)
    buf[0:name_size] = name_utf16
    buf[64:66] = struct.pack("<H", name_size)
    buf[66:67] = struct.pack("B", entry["type"])  # object type
    buf[67:68] = struct.pack("B", entry.get("color", 1))  # color (1=black)
    buf[68:72] = struct.pack("<I", entry.get("left_sibling", nostream))
    buf[72:76] = struct.pack("<I", entry.get("right_sibling", nostream))
    buf[76:80] = struct.pack("<I", entry.get("child_id", nostream))

    # CLSID (zeros for our purposes)
    # State bits (zeros)

    if entry["type"] == 5:
        # Root entry: start = first sector of mini-stream
        buf[116:120] = struct.pack("<I", entry.get("mini_start", 0))
    elif entry["type"] == 2:
        # Stream: start = first mini-sector
        buf[116:120] = struct.pack("<I", entry.get("mini_start", nostream))
    else:
        buf[116:120] = struct.pack("<I", 0)

    buf[120:124] = struct.pack("<I", entry.get("size", 0))

    return bytes(buf)


def _build_cfb_header(
    total_sectors: int,
    fat_sectors: list[int],
    dir_start: int,
    mini_fat_start: int,
    n_mini_fat_sectors: int,
) -> bytes:
    """Build the 512-byte CFB header."""
    header = bytearray(512)

    # Signature
    header[0:8] = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
    # Minor version
    header[24:26] = struct.pack("<H", 0x003E)
    # Major version (3 = v3)
    header[26:28] = struct.pack("<H", 0x0003)
    # Byte order (little-endian)
    header[28:30] = struct.pack("<H", 0xFFFE)
    # Sector size power (9 = 512)
    header[30:32] = struct.pack("<H", 9)
    # Mini sector size power (6 = 64)
    header[32:34] = struct.pack("<H", 6)
    # Total directory sectors (0 for v3)
    header[40:44] = struct.pack("<I", 0)
    # Total FAT sectors
    header[44:48] = struct.pack("<I", len(fat_sectors))
    # First directory sector
    header[48:52] = struct.pack("<I", dir_start)
    # Transaction signature (0)
    header[52:56] = struct.pack("<I", 0)
    # Mini stream cutoff (4096)
    header[56:60] = struct.pack("<I", 0x1000)
    # First mini-FAT sector
    header[60:64] = struct.pack("<I", mini_fat_start)
    # Total mini-FAT sectors
    header[64:68] = struct.pack("<I", n_mini_fat_sectors)
    # First DIFAT sector (none)
    header[68:72] = struct.pack("<I", 0xFFFFFFFE)
    # Total DIFAT sectors
    header[72:76] = struct.pack("<I", 0)
    # DIFAT array (109 entries)
    for i in range(109):
        offset = 76 + i * 4
        if i < len(fat_sectors):
            header[offset:offset + 4] = struct.pack("<I", fat_sectors[i])
        else:
            header[offset:offset + 4] = struct.pack("<I", 0xFFFFFFFF)

    return bytes(header)


# ============================================================================
# ACTUARIAL WORKBOOK BUILDER
# ============================================================================

# VBA Module source code
VBA_MODULES = {
    "ActuarialEngine": r'''Option Explicit

'===================================================================
' ACTUARIAL ENGINE MODULE
' Insurance calculation routines with Chain Ladder, BF, and SCR
' Author: Actuarial Department  |  Version: 3.0
'===================================================================

'-------------------------------------------------------------------
' SUB 1: MASTER RUNNER - Executes complete actuarial model refresh
'-------------------------------------------------------------------
Sub RunFullActuarialModel()
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationAutomatic
    
    Call RefreshAssumptions
    Call UpdateMortalityRates
    Call CalculatePremiums
    Call UpdateChainLadderFactors
    Call CalculateBornhuetterFerguson
    Call ComputeLossRatios
    Call UpdateSolvencyCapital
    Call RefreshDashboard
    
    Application.ScreenUpdating = True
    MsgBox "Actuarial model refreshed: " & Format(Date, "DD-MMM-YYYY"), _
           vbInformation, "Actuarial Suite"
End Sub

'-------------------------------------------------------------------
' SUB 2: REFRESH ASSUMPTIONS
'-------------------------------------------------------------------
Sub RefreshAssumptions()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("Assumptions")
    ws.Calculate
End Sub

'-------------------------------------------------------------------
' SUB 3: UPDATE MORTALITY RATES (Makeham model)
'-------------------------------------------------------------------
Sub UpdateMortalityRates()
    Dim ws As Worksheet, wsA As Worksheet
    Dim age As Long
    Dim baseRate As Double, improvFactor As Double, discRate As Double
    
    Set ws = ThisWorkbook.Sheets("MortalityTable")
    Set wsA = ThisWorkbook.Sheets("Assumptions")
    
    baseRate = wsA.Range("C18").Value
    improvFactor = wsA.Range("C19").Value
    discRate = wsA.Range("C21").Value
    
    For age = 20 To 80
        Dim row As Long
        row = age - 16
        Dim qx As Double
        qx = baseRate * (1 - improvFactor) ^ (age - 20)
        
        ws.Cells(row, 3).Value = age
        ws.Cells(row, 4).Value = qx
        ws.Cells(row, 5).Value = 1 - qx
        ws.Cells(row, 6).Value = qx * 1000
        ws.Cells(row, 7).Value = (1 - qx / (discRate + qx)) / Application.WorksheetFunction.Ln(1 + discRate)
        ws.Cells(row, 7).NumberFormat = "0.0000"
    Next age
End Sub

'-------------------------------------------------------------------
' SUB 4: CALCULATE PREMIUMS (Gross & Net)
'-------------------------------------------------------------------
Sub CalculatePremiums()
    Dim ws As Worksheet, wsA As Worksheet
    Dim i As Long
    Dim sumInsured As Double, lives As Double
    Dim grossPrem As Double
    Dim expenseLoad As Double, profitLoad As Double, reinsLoad As Double
    
    Set ws = ThisWorkbook.Sheets("Premiums")
    Set wsA = ThisWorkbook.Sheets("Assumptions")
    
    expenseLoad = wsA.Range("C6").Value
    profitLoad = wsA.Range("C7").Value
    reinsLoad = wsA.Range("C8").Value
    
    For i = 5 To 14
        lives = ws.Cells(i, 4).Value
        sumInsured = ws.Cells(i, 5).Value
        grossPrem = lives * sumInsured
        
        ws.Cells(i, 6).Value = grossPrem
        ws.Cells(i, 7).Value = expenseLoad
        ws.Cells(i, 8).Value = profitLoad
        ws.Cells(i, 9).Value = reinsLoad
        ws.Cells(i, 10).Value = grossPrem / (1 - expenseLoad - profitLoad - reinsLoad)
        ws.Cells(i, 10).NumberFormat = "#,##0"
    Next i
    
    ' Totals
    ws.Range("F20").Value = Application.WorksheetFunction.Sum(ws.Range("F5:F14"))
    ws.Range("J20").Value = Application.WorksheetFunction.Sum(ws.Range("J5:J14"))
    ws.Range("F21").Value = ws.Range("J20").Value
    ws.Range("F22").Value = ws.Range("F21").Value / ws.Range("F20").Value
End Sub

'-------------------------------------------------------------------
' SUB 5: CHAIN LADDER - WEIGHTED-AVERAGE DEVELOPMENT FACTORS
'-------------------------------------------------------------------
Sub UpdateChainLadderFactors()
    Dim ws As Worksheet, wsA As Worksheet
    Dim sumN As Double, sumD As Double
    Dim i As Long, j As Long
    
    Set ws = ThisWorkbook.Sheets("ChainLadder")
    Set wsA = ThisWorkbook.Sheets("Assumptions")
    
    For j = 3 To 11
        sumN = 0: sumD = 0
        For i = 4 To 13
            If ws.Cells(i, j).Value > 0 And ws.Cells(i, j + 1).Value > 0 Then
                sumN = sumN + ws.Cells(i, j + 1).Value
                sumD = sumD + ws.Cells(i, j).Value
            End If
        Next i
        If sumD > 0 Then
            ws.Cells(15, j).Value = sumN / sumD
            ws.Cells(15, j).NumberFormat = "0.0000"
        End If
    Next j
    
    ' Cumulative factors
    Dim cumProd As Double
    cumProd = 1
    For j = 11 To 3 Step -1
        cumProd = cumProd * ws.Cells(15, j).Value
        ws.Cells(16, j).Value = cumProd
        ws.Cells(16, j).NumberFormat = "0.0000"
    Next j
    ws.Cells(15, 12).Value = wsA.Range("C23").Value
End Sub

'-------------------------------------------------------------------
' SUB 6: BORNHUETTER-FERGUSON IBNR CALCULATION
'-------------------------------------------------------------------
Sub CalculateBornhuetterFerguson()
    Dim ws As Worksheet, wsA As Worksheet
    Dim BF_LR As Double
    Dim aPriori As Double, pctDev As Double
    Dim bfIBNR As Double, clIBNR As Double
    Dim i As Long
    
    Set ws = ThisWorkbook.Sheets("IBNR")
    Set wsA = ThisWorkbook.Sheets("Assumptions")
    BF_LR = wsA.Range("C24").Value
    
    For i = 4 To 13
        aPriori = ws.Cells(i, 4).Value
        pctDev = ws.Cells(i, 5).Value
        clIBNR = ws.Cells(i, 6).Value
        bfIBNR = aPriori * BF_LR * (1 - pctDev)
        
        ws.Cells(i, 7).Value = bfIBNR
        ws.Cells(i, 7).NumberFormat = "#,##0"
        
        If pctDev < 0.5 Then
            ws.Cells(i, 8).Value = bfIBNR
            ws.Cells(i, 11).Value = "BF"
        Else
            ws.Cells(i, 8).Value = clIBNR
            ws.Cells(i, 11).Value = "CL"
        End If
        ws.Cells(i, 8).NumberFormat = "#,##0"
    Next i
    
    ws.Range("H14").FormulaR1C1 = "=SUM(R4C8:R13C8)"
End Sub

'-------------------------------------------------------------------
' SUB 7: COMPUTE LOSS RATIOS WITH TRAFFIC LIGHTS
'-------------------------------------------------------------------
Sub ComputeLossRatios()
    Dim ws As Worksheet
    Dim cell As Range
    Dim lr As Double
    
    Set ws = ThisWorkbook.Sheets("LossRatios")
    
    For Each cell In ws.Range("G4:G13")
        If IsNumeric(cell.Value) And cell.Value > 0 Then
            lr = cell.Value
            Select Case True
                Case lr < 0.55
                    cell.Interior.Color = RGB(0, 176, 80)
                    cell.Font.Color = vbWhite
                Case lr < 0.65
                    cell.Interior.Color = RGB(255, 255, 0)
                    cell.Font.Color = vbBlack
                Case lr < 0.75
                    cell.Interior.Color = RGB(255, 192, 0)
                    cell.Font.Color = vbBlack
                Case Else
                    cell.Interior.Color = RGB(255, 0, 0)
                    cell.Font.Color = vbWhite
            End Select
        End If
    Next cell
End Sub

'-------------------------------------------------------------------
' SUB 8: SOLVENCY CAPITAL REQUIREMENT
'-------------------------------------------------------------------
Sub UpdateSolvencyCapital()
    Dim ws As Worksheet
    Dim ratio As Double
    
    Set ws = ThisWorkbook.Sheets("Solvency")
    ws.Calculate
    ratio = ws.Range("C32").Value
    
    Select Case True
        Case ratio >= 1.5
            ws.Range("C32").Interior.Color = RGB(0, 176, 80)
            ws.Range("C32").Font.Color = vbWhite
        Case ratio >= 1.0
            ws.Range("C32").Interior.Color = RGB(255, 255, 0)
            ws.Range("C32").Font.Color = vbBlack
        Case Else
            ws.Range("C32").Interior.Color = RGB(255, 0, 0)
            ws.Range("C32").Font.Color = vbWhite
            MsgBox "CRITICAL: SCR Coverage = " & Format(ratio, "0.0x") & _
                   " (BREACH!)", vbCritical, "Solvency Breach"
    End Select
End Sub

'-------------------------------------------------------------------
' SUB 9: REFRESH DASHBOARD
'-------------------------------------------------------------------
Sub RefreshDashboard()
    Dim wsDash As Worksheet
    Set wsDash = ThisWorkbook.Sheets("Dashboard")
    wsDash.Calculate
End Sub

'-------------------------------------------------------------------
' SUB 10: STRESS TEST - 1-IN-200 YEAR SCENARIO
'-------------------------------------------------------------------
Sub RunStressScenario()
    Dim wsAss As Worksheet, wsCL As Worksheet, wsSol As Worksheet
    Dim origLR As Double, origCat As Double
    Dim baseIBNR As Double, stressIBNR As Double
    Dim baseSCR As Double, stressSCR As Double
    
    Set wsAss = ThisWorkbook.Sheets("Assumptions")
    Set wsCL = ThisWorkbook.Sheets("ChainLadder")
    Set wsSol = ThisWorkbook.Sheets("Solvency")
    
    origLR = wsAss.Range("C11").Value
    origCat = wsSol.Range("C20").Value
    baseIBNR = wsCL.Range("M20").Value
    baseSCR = wsSol.Range("C28").Value
    
    ' Apply 1-in-200 stresses
    wsAss.Range("C11").Value = origLR * 1.35
    wsSol.Range("C20").Value = origCat * 3#
    Application.Calculate
    
    stressIBNR = wsCL.Range("M20").Value
    stressSCR = wsSol.Range("C28").Value
    
    Dim msg As String
    msg = "=== 1-IN-200 STRESS SCENARIO ==" & vbCrLf & vbCrLf
    msg = msg & "IBNR Impact:  " & Format(stressIBNR - baseIBNR, "#,##0") & vbCrLf
    msg = msg & "SCR Impact:   " & Format(stressSCR - baseSCR, "#,##0") & vbCrLf
    msg = msg & "New Coverage: " & Format(wsSol.Range("C32").Value, "0.0x")
    MsgBox msg, vbExclamation, "Stress Test Result"
    
    ' Restore
    wsAss.Range("C11").Value = origLR
    wsSol.Range("C20").Value = origCat
    Application.Calculate
End Sub

'-------------------------------------------------------------------
' SUB 11: VALIDATE MODEL
'-------------------------------------------------------------------
Sub ValidateModel()
    Dim wsAss As Worksheet
    Dim issues As String
    Set wsAss = ThisWorkbook.Sheets("Assumptions")
    issues = ""
    
    If wsAss.Range("C11").Value > 0.95 Or wsAss.Range("C11").Value < 0.3 Then
        issues = issues & "[WARN] Loss ratio outside 30-95% range" & vbCrLf
    End If
    If wsAss.Range("C4").Value < 0 Then
        issues = issues & "[ERROR] Risk-free rate is negative" & vbCrLf
    End If
    If wsAss.Range("C23").Value < 1# Then
        issues = issues & "[ERROR] Tail factor < 1.0 is invalid" & vbCrLf
    End If
    
    If issues = "" Then
        MsgBox "All validation checks passed.", vbInformation, "OK"
    Else
        MsgBox "ISSUES:" & vbCrLf & issues, vbExclamation, "Warnings"
    End If
End Sub
''',
}


def _add_assumptions_sheet(wb: openpyxl.Workbook) -> None:
    """Populate the Assumptions sheet with actuarial parameters."""
    ws = wb.create_sheet("Assumptions")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    label_font = Font(bold=True, size=10)
    num_fmt = "#,##0.0000"
    pct_fmt = "0.00%"

    # Title
    ws.merge_cells("B1:D1")
    ws["B1"] = "ACTUARIAL ASSUMPTIONS"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    assumptions = [
        ("", "", "", ""),
        ("General", "", "", ""),
        ("Risk-Free Rate", 0.035, pct_fmt, "Solvency II curve"),
        ("Inflation Rate", 0.025, pct_fmt, "CPI assumption"),
        ("Expense Loading", 0.12, pct_fmt, "% of gross premium"),
        ("Profit Loading", 0.05, pct_fmt, "Target profit margin"),
        ("Reinsurance Loading", 0.08, pct_fmt, "QS cession rate"),
        ("", "", "", ""),
        ("Claims", "", "", ""),
        ("Expected Loss Ratio", 0.62, pct_fmt, "A priori loss ratio"),
        ("Claims Inflation", 0.04, pct_fmt, "Superimposed inflation"),
        ("", "", "", ""),
        ("Premium", "", "", ""),
        ("Rate Change", 0.03, pct_fmt, "Renewal adjustment"),
        ("Retention Rate", 0.88, pct_fmt, "Policy retention"),
        ("Average IBNR Factor", 1.15, num_fmt, "IBNR/case ratio"),
        ("", "", "", ""),
        ("Mortality", "", "", ""),
        ("Base Mortality Rate (q20)", 0.001, "0.00000", "Age 20 base"),
        ("Improvement Factor", 0.015, pct_fmt, "Annual improvement"),
        ("Mortality Loading", 0.10, pct_fmt, "Safety margin"),
        ("Discount Rate", 0.04, pct_fmt, "For annuity values"),
        ("", "", "", ""),
        ("Development", "", "", ""),
        ("Tail Factor", 1.02, num_fmt, "Beyond triangle"),
        ("BF A Priori LR", 0.65, pct_fmt, "BF expected LR"),
        ("BF Confidence Weight", 0.60, pct_fmt, "BF vs CL blend"),
        ("", "", "", ""),
        ("Solvency", "", "", ""),
        ("SCR Premium Factor", 0.09, pct_fmt, "Premium risk"),
        ("SCR Reserve Factor", 0.11, pct_fmt, "Reserve risk"),
        ("Diversification Benefit", 0.20, pct_fmt, "Correlation adj"),
        ("Cat Risk Load", 5000, "#,##0", "£000s"),
        ("Available Capital", 35000, "#,##0", "£000s"),
    ]

    for i, (label, value, fmt, note) in enumerate(assumptions, start=2):
        ws.cell(row=i, column=2, value=label).font = label_font
        if value != "":
            c = ws.cell(row=i, column=3, value=value)
            if fmt:
                c.number_format = fmt
        if note:
            ws.cell(row=i, column=4, value=note).font = Font(italic=True, color="666666")

    # Section headers
    for row_label in ["General", "Claims", "Premium", "Mortality", "Development", "Solvency"]:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                if cell.value == row_label:
                    cell.font = Font(bold=True, size=11, color="1F3864")
                    cell.fill = PatternFill("solid", fgColor="D6E4F0")

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22


def _add_mortality_sheet(wb: openpyxl.Workbook) -> None:
    """Populate MortalityTable with age-based mortality data."""
    ws = wb.create_sheet("MortalityTable")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:I1")
    ws["B1"] = "MORTALITY TABLE - Makeham Model"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["#", "Age", "qx", "px (=1-qx)", "lx", "dx (=lx*qx)", "1000qx", "äx"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    base_qx = 0.001
    improvement = 0.015
    disc = 0.04
    lx = 100000

    for i, age in enumerate(range(20, 81)):
        row = i + 4
        qx = base_qx * (1 - improvement) ** (age - 20)
        px = 1 - qx
        dx = lx * qx

        ws.cell(row=row, column=2, value=i + 1)
        ws.cell(row=row, column=3, value=age)
        ws.cell(row=row, column=4, value=round(qx, 6)).number_format = "0.000000"
        ws.cell(row=row, column=5, value=round(px, 6)).number_format = "0.000000"
        ws.cell(row=row, column=6, value=round(lx, 0)).number_format = "#,##0"
        ws.cell(row=row, column=7, value=round(dx, 2)).number_format = "#,##0.00"
        ws.cell(row=row, column=8, value=round(qx * 1000, 4)).number_format = "0.0000"

        ax = (1 - qx / (disc + qx)) / np.log(1 + disc)
        ws.cell(row=row, column=9, value=round(ax, 4)).number_format = "0.0000"

        lx = lx * px

    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14


def _add_premiums_sheet(wb: openpyxl.Workbook) -> None:
    """Populate Premiums sheet with policy and premium data."""
    ws = wb.create_sheet("Premiums")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:J1")
    ws["B1"] = "PREMIUM CALCULATION"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["Year", "Line of Business", "Lives", "Sum Insured (£k)",
               "Gross Premium", "Expense %", "Profit %", "RI %", "Net Premium"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    lobs = [
        ("Motor", 2500, 8.5), ("Property", 1800, 12.0), ("Liability", 950, 18.5),
        ("Marine", 420, 25.0), ("Accident", 3200, 5.2), ("Health", 1500, 7.8),
        ("Life", 800, 45.0), ("Pension", 650, 32.0), ("Travel", 5000, 2.1),
        ("Specialty", 300, 55.0),
    ]

    expense, profit, ri = 0.12, 0.05, 0.08
    total_gross = 0
    total_net = 0

    for i, (lob, lives, si) in enumerate(lobs):
        row = i + 5  # data starts row 5 (after blank row 4 as spacer)
        year = 2024 + (i % 3)
        gross = lives * si
        net = gross / (1 - expense - profit - ri)
        total_gross += gross
        total_net += net

        ws.cell(row=row, column=2, value=year)
        ws.cell(row=row, column=3, value=lob)
        ws.cell(row=row, column=4, value=lives).number_format = "#,##0"
        ws.cell(row=row, column=5, value=si).number_format = "#,##0.0"
        ws.cell(row=row, column=6, value=round(gross, 0)).number_format = "#,##0"
        ws.cell(row=row, column=7, value=expense).number_format = "0.00%"
        ws.cell(row=row, column=8, value=profit).number_format = "0.00%"
        ws.cell(row=row, column=9, value=ri).number_format = "0.00%"
        ws.cell(row=row, column=10, value=round(net, 0)).number_format = "#,##0"

    # Totals
    ws.cell(row=20, column=2, value="TOTAL GWP").font = Font(bold=True)
    ws.cell(row=20, column=6, value=round(total_gross, 0)).number_format = "#,##0"
    ws.cell(row=20, column=10, value=round(total_net, 0)).number_format = "#,##0"
    ws.cell(row=21, column=2, value="NET WRITTEN").font = Font(bold=True)
    ws.cell(row=21, column=6, value=round(total_net, 0)).number_format = "#,##0"
    ws.cell(row=22, column=2, value="COMBINED RATIO").font = Font(bold=True)
    ws.cell(row=22, column=6, value=round(total_net / total_gross, 4)).number_format = "0.00%"

    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _add_chain_ladder_sheet(wb: openpyxl.Workbook) -> None:
    """Populate ChainLadder with claims triangle and development factors."""
    ws = wb.create_sheet("ChainLadder")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:M1")
    ws["B1"] = "CHAIN LADDER CLAIMS TRIANGLE"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    # Headers: Accident Year, Dev 1..10, Ultimate
    dev_headers = ["AY"] + [f"Dev {d}" for d in range(1, 11)] + ["Ultimate"]
    for j, h in enumerate(dev_headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")

    # Generate synthetic claims triangle
    np.random.seed(42)
    base_claims = np.array([4500, 5200, 4800, 5500, 6100, 5900, 6500, 7200, 6800, 7500], dtype=float)
    dev_factors = [2.5, 1.6, 1.3, 1.15, 1.08, 1.05, 1.03, 1.02, 1.01, 1.005]

    triangle = np.zeros((10, 11))
    for i in range(10):
        triangle[i, 0] = base_claims[i] * (1 + np.random.normal(0, 0.05))
        for j in range(1, 11 - i):
            noise = 1 + np.random.normal(0, 0.02)
            triangle[i, j] = triangle[i, j - 1] * dev_factors[j - 1] * noise

    for i in range(10):
        ws.cell(row=i + 4, column=2, value=2015 + i)
        for j in range(11 - i):
            ws.cell(row=i + 4, column=j + 3, value=round(triangle[i, j], 0)).number_format = "#,##0"

    # Dev factor row
    ws.cell(row=15, column=2, value="LDF").font = Font(bold=True)
    for j in range(10):
        num = sum(triangle[i, j + 1] for i in range(10 - j - 1) if triangle[i, j + 1] > 0)
        den = sum(triangle[i, j] for i in range(10 - j - 1) if triangle[i, j + 1] > 0)
        ldf = num / den if den > 0 else 1.0
        ws.cell(row=15, column=j + 3, value=round(ldf, 4)).number_format = "0.0000"

    # Cumulative
    ws.cell(row=16, column=2, value="CDF").font = Font(bold=True)
    cdf = 1.0
    cdfs = []
    for j in range(9, -1, -1):
        ldf_val = ws.cell(row=15, column=j + 3).value or 1.0
        cdf *= ldf_val
        cdfs.insert(0, cdf)
    for j, c in enumerate(cdfs):
        ws.cell(row=16, column=j + 3, value=round(c, 4)).number_format = "0.0000"

    # Ultimate column
    tail = 1.02
    ws.cell(row=15, column=13, value=tail).number_format = "0.0000"
    for i in range(10):
        last_dev = 10 - i
        last_val = ws.cell(row=i + 4, column=last_dev + 2).value
        if last_val and last_dev - 1 < len(cdfs):
            ult = last_val * (cdfs[last_dev - 1] if last_dev - 1 < len(cdfs) else 1) * tail
            ws.cell(row=i + 4, column=13, value=round(ult, 0)).number_format = "#,##0"

    ws.cell(row=20, column=2, value="TOTAL ULTIMATE").font = Font(bold=True, size=11)
    total_ult = sum(ws.cell(row=i + 4, column=13).value or 0 for i in range(10))
    ws.cell(row=20, column=13, value=round(total_ult, 0)).number_format = "#,##0"

    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13


def _add_ibnr_sheet(wb: openpyxl.Workbook) -> None:
    """Populate IBNR sheet with BF and CL reserve estimates."""
    ws = wb.create_sheet("IBNR")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:K1")
    ws["B1"] = "IBNR RESERVE ESTIMATES"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["AY", "Paid to Date", "Projected Ult (CL)",
               "A Priori EP", "% Developed", "CL IBNR",
               "BF IBNR", "Selected IBNR", "Case Reserve",
               "Method"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    np.random.seed(123)
    bf_lr = 0.65
    for i in range(10):
        row = i + 4
        ay = 2015 + i
        paid = round(np.random.uniform(3000, 9000) * (10 - i) / 10, 0)
        pct_dev = round(min(0.95, (10 - i) / 10 + np.random.uniform(-0.05, 0.05)), 4)
        a_priori = round(paid / max(pct_dev, 0.1), 0)
        projected = round(paid / max(pct_dev, 0.1), 0)
        cl_ibnr = round(projected - paid, 0)
        bf_ibnr = round(a_priori * bf_lr * (1 - pct_dev), 0)
        selected = bf_ibnr if pct_dev < 0.5 else cl_ibnr
        case_reserve = round(selected * 0.8, 0)
        method = "BF" if pct_dev < 0.5 else "CL"

        ws.cell(row=row, column=2, value=ay)
        ws.cell(row=row, column=3, value=paid).number_format = "#,##0"
        ws.cell(row=row, column=4, value=projected).number_format = "#,##0"
        ws.cell(row=row, column=5, value=a_priori).number_format = "#,##0"
        ws.cell(row=row, column=6, value=pct_dev).number_format = "0.00%"
        ws.cell(row=row, column=7, value=cl_ibnr).number_format = "#,##0"
        ws.cell(row=row, column=8, value=bf_ibnr).number_format = "#,##0"
        ws.cell(row=row, column=9, value=selected).number_format = "#,##0"
        ws.cell(row=row, column=10, value=case_reserve).number_format = "#,##0"
        ws.cell(row=row, column=11, value=method)

        # Color: BF = amber, CL = blue
        fill = PatternFill("solid", fgColor="FFE699" if method == "BF" else "BDD7EE")
        ws.cell(row=row, column=9).fill = fill

    ws.cell(row=14, column=2, value="TOTAL").font = Font(bold=True)
    for col in [7, 8, 9, 10]:
        total = sum(ws.cell(row=r, column=col).value or 0 for r in range(4, 14))
        ws.cell(row=14, column=col, value=round(total, 0)).number_format = "#,##0"
        ws.cell(row=14, column=col).font = Font(bold=True)

    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _add_loss_ratios_sheet(wb: openpyxl.Workbook) -> None:
    """Populate LossRatios sheet."""
    ws = wb.create_sheet("LossRatios")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:K1")
    ws["B1"] = "LOSS RATIO ANALYSIS"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    headers = ["AY", "Earned Premium", "Paid Claims",
               "IBNR (CL)", "IBNR (Selected)", "Incurred",
               "Loss Ratio", "Normalised LR", "Combined Ratio",
               "Status"]
    for j, h in enumerate(headers, start=2):
        c = ws.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    np.random.seed(456)
    for i in range(10):
        row = i + 4
        ay = 2015 + i
        ep = round(np.random.uniform(8000, 15000), 0)
        paid = round(ep * np.random.uniform(0.35, 0.55), 0)
        ibnr_cl = round(ep * np.random.uniform(0.05, 0.20), 0)
        ibnr_sel = round(ibnr_cl * np.random.uniform(0.8, 1.1), 0)
        incurred = paid + ibnr_sel
        lr = incurred / ep if ep > 0 else 0
        norm_lr = lr / 1.0  # normalised
        combined = lr + 0.12 + 0.05  # + expense + profit load
        status = "Green" if lr < 0.55 else ("Amber" if lr < 0.65 else ("Watch" if lr < 0.75 else "Red"))

        ws.cell(row=row, column=2, value=ay)
        ws.cell(row=row, column=3, value=ep).number_format = "#,##0"
        ws.cell(row=row, column=4, value=paid).number_format = "#,##0"
        ws.cell(row=row, column=5, value=ibnr_cl).number_format = "#,##0"
        ws.cell(row=row, column=6, value=ibnr_sel).number_format = "#,##0"
        ws.cell(row=row, column=7, value=incurred).number_format = "#,##0"
        c = ws.cell(row=row, column=8, value=round(lr, 4))
        c.number_format = "0.00%"
        # Traffic light
        if lr < 0.55:
            c.fill = PatternFill("solid", fgColor="00B050")
            c.font = Font(color="FFFFFF")
        elif lr < 0.65:
            c.fill = PatternFill("solid", fgColor="FFFF00")
        elif lr < 0.75:
            c.fill = PatternFill("solid", fgColor="FFC000")
        else:
            c.fill = PatternFill("solid", fgColor="FF0000")
            c.font = Font(color="FFFFFF")

        ws.cell(row=row, column=9, value=round(norm_lr, 4)).number_format = "0.00%"
        ws.cell(row=row, column=10, value=round(combined, 4)).number_format = "0.00%"
        ws.cell(row=row, column=11, value=status)

    for col in range(2, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _add_solvency_sheet(wb: openpyxl.Workbook) -> None:
    """Populate Solvency sheet with SCR calculation."""
    ws = wb.create_sheet("Solvency")
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("B1:D1")
    ws["B1"] = "SOLVENCY CAPITAL REQUIREMENT"
    ws["B1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    rows_data = [
        ("", "", ""),
        ("PREMIUM RISK", "", ""),
        ("Net Written Premium", 28400, "#,##0"),
        ("SCR Premium Factor", 0.09, "0.00%"),
        ("Premium Risk Capital", 2556, "#,##0"),
        ("", "", ""),
        ("RESERVE RISK", "", ""),
        ("Total IBNR Reserve", 12500, "#,##0"),
        ("SCR Reserve Factor", 0.11, "0.00%"),
        ("Reserve Risk Capital", 1375, "#,##0"),
        ("", "", ""),
        ("UNDERWRITING RISK", "", ""),
        ("Gross Underwriting Risk", 3931, "#,##0"),
        ("Diversification Benefit", -786, "#,##0"),
        ("Net Underwriting Risk", 3145, "#,##0"),
        ("", "", ""),
        ("MARKET RISK", "", ""),
        ("Interest Rate Risk", 1200, "#,##0"),
        ("Equity Risk", 800, "#,##0"),
        ("Catastrophe Risk", 5000, "#,##0"),
        ("Total Market Risk", 7000, "#,##0"),
        ("", "", ""),
        ("OPERATIONAL RISK", "", ""),
        ("Op Risk (3% of premium)", 852, "#,##0"),
        ("", "", ""),
        ("TOTAL SCR", "", ""),
        ("Total SCR", 10997, "#,##0"),
        ("Adjustment", -500, "#,##0"),
        ("SCR After Adjustment", 10497, "#,##0"),
        ("", "", ""),
        ("Available Capital", 35000, "#,##0"),
        ("SCR Coverage Ratio", 3.334, "0.00x"),
    ]

    for i, (label, value, fmt) in enumerate(rows_data, start=2):
        c = ws.cell(row=i, column=2, value=label)
        if label and value == "" and fmt == "":
            c.font = Font(bold=True, size=11, color="1F3864")
            c.fill = PatternFill("solid", fgColor="D6E4F0")
        else:
            c.font = Font(bold=True) if "TOTAL" in label or "Coverage" in label else Font()

        if value != "":
            v = ws.cell(row=i, column=3, value=value)
            if fmt:
                v.number_format = fmt

    # Color the coverage ratio
    ratio_row = 2 + len(rows_data) - 1
    coverage_cell = ws.cell(row=ratio_row, column=3)
    coverage_cell.fill = PatternFill("solid", fgColor="00B050")
    coverage_cell.font = Font(bold=True, color="FFFFFF", size=12)

    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _add_dashboard_sheet(wb: openpyxl.Workbook) -> None:
    """Create the Dashboard sheet with summary KPIs."""
    ws = wb.create_sheet("Dashboard", 0)  # Insert as first sheet
    header_fill = PatternFill("solid", fgColor="1F3864")

    ws.merge_cells("B1:G1")
    ws["B1"] = "ACTUARIAL DASHBOARD"
    ws["B1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["B1"].fill = header_fill
    ws["B1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:G2")
    ws["B2"] = f"As at {__import__('datetime').date.today().strftime('%d %B %Y')}"
    ws["B2"].font = Font(size=11, color="666666")
    ws["B2"].alignment = Alignment(horizontal="center")

    kpis = [
        ("", "", "", "", "", ""),
        ("KPI", "Value", "Target", "Variance", "RAG", ""),
        ("", "", "", "", "", ""),
        ("Gross Written Premium (£k)", 171350, 165000, 6350, "Green", ""),
        ("Combined Ratio", "133.33%", "< 100%", "", "Red", ""),
        ("Total IBNR Reserve (£k)", 12500, 15000, -2500, "Green", ""),
        ("SCR Coverage", "3.33x", "> 1.50x", "", "Green", ""),
        ("Ultimate Claims (£k)", 85000, 90000, -5000, "Green", ""),
        ("Loss Ratio (Current)", "62.0%", "< 65%", "", "Green", ""),
        ("Reserve Adequacy", "103.5%", "> 100%", "", "Green", ""),
        ("Expense Ratio", "12.0%", "< 15%", "", "Green", ""),
    ]

    for i, (label, val, target, var, rag, _) in enumerate(kpis, start=4):
        ws.cell(row=i, column=2, value=label)
        ws.cell(row=i, column=3, value=val)
        ws.cell(row=i, column=4, value=target)
        ws.cell(row=i, column=5, value=var if var != "" else "")
        ws.cell(row=i, column=6, value=rag)

        if label == "KPI":
            for col in range(2, 7):
                ws.cell(row=i, column=col).font = Font(bold=True, color="FFFFFF")
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="4472C4")
        elif rag:
            color_map = {"Green": "00B050", "Amber": "FFC000", "Red": "FF0000"}
            rag_cell = ws.cell(row=i, column=6)
            if rag in color_map:
                rag_cell.fill = PatternFill("solid", fgColor=color_map[rag])
                if rag != "Amber":
                    rag_cell.font = Font(color="FFFFFF", bold=True)

    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 22


def create_actuarial_workbook(output_path: str) -> str:
    """
    Create the complete .xlsm workbook with data and embedded VBA.

    Args:
        output_path: Where to save the file.

    Returns:
        Absolute path to the created file.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets
    _add_dashboard_sheet(wb)
    _add_assumptions_sheet(wb)
    _add_mortality_sheet(wb)
    _add_premiums_sheet(wb)
    _add_chain_ladder_sheet(wb)
    _add_ibnr_sheet(wb)
    _add_loss_ratios_sheet(wb)
    _add_solvency_sheet(wb)

    # Build the VBA project binary
    print("Building vbaProject.bin with embedded VBA macros...")
    vba_bin = build_vba_project_bin("ActuarialProject", VBA_MODULES)
    print(f"  vbaProject.bin size: {len(vba_bin):,} bytes")

    # Save as xlsx first, then inject VBA to make it xlsm
    import tempfile
    import zipfile
    import shutil

    tmp_xlsx = tempfile.mktemp(suffix=".xlsx")
    wb.save(tmp_xlsx)
    wb.close()

    # Convert to xlsm by injecting vbaProject.bin into the ZIP
    with zipfile.ZipFile(tmp_xlsx, "r") as zin:
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "[Content_Types].xml":
                    # Add VBA content type
                    data = data.replace(
                        b"</Types>",
                        b'<Override PartName="/xl/vbaProject.bin" '
                        b'ContentType="application/vnd.ms-office.vbaProject"/>'
                        b"\n</Types>",
                    )
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    # Add relationship to vbaProject.bin
                    data = data.replace(
                        b"</Relationships>",
                        b'<Relationship Id="rIdVBA" '
                        b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                        b'Target="vbaProject.bin"/>'
                        b"\n</Relationships>",
                    )
                zout.writestr(item, data)

            # Add the VBA binary
            zout.writestr("xl/vbaProject.bin", vba_bin)

    os.unlink(tmp_xlsx)
    abs_path = os.path.abspath(output_path)
    print(f"\nCreated: {abs_path}")
    print(f"File size: {os.path.getsize(abs_path):,} bytes")
    return abs_path


def main() -> None:
    default_name = "Actuarial_Calculations.xlsm"
    output = sys.argv[1] if len(sys.argv) > 1 else default_name
    path = create_actuarial_workbook(output)

    # Verify with oletools
    print("\n--- Verification with oletools ---")
    try:
        from oletools.olevba import VBA_Parser

        vba = VBA_Parser(path)
        if vba.detect_vba_macros():
            print("✓ VBA macros DETECTED")
            for filename, stream, vba_name, vba_code in vba.extract_macros():
                lines = len(vba_code.strip().splitlines()) if vba_code else 0
                print(f"  Module: {vba_name}  ({lines} lines)")
        else:
            print("✗ No VBA macros detected (binary may need refinement)")
        vba.close()
    except Exception as e:
        print(f"  oletools check failed: {e}")

    print(f"\nDone! File ready at: {path}")


if __name__ == "__main__":
    main()
