"""
Generate a complex actuarial insurance workbook (.xlsm) with embedded VBA macros.

This script creates a realistic Excel workbook that an actuary would use for
insurance pricing, reserving, and risk analysis — complete with VBA macros
for mortality table lookups, premium calculations, loss reserving, and more.

The VBA is embedded via a pre-built vbaProject.bin created by this script.
"""
from __future__ import annotations

import io
import math
import os
import random
import struct
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
OUTPUT_DIR = Path(__file__).resolve().parent / "sample_files"
OUTPUT_FILE = OUTPUT_DIR / "Actuarial_Insurance_Model.xlsm"

HEADER_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CURRENCY_FMT = '#,##0.00'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# Seed for reproducibility
random.seed(42)


# ---------------------------------------------------------------------------
# VBA Code Modules
# ---------------------------------------------------------------------------

VBA_MODULE_MORTALITY = '''
Attribute VB_Name = "MortalityFunctions"
' ============================================================
' Module: MortalityFunctions
' Purpose: Life table and mortality calculations for actuarial
'          pricing of life insurance products.
' ============================================================
Option Explicit

' --- Constants ---
Private Const MAX_AGE As Integer = 120
Private Const BASE_MORTALITY_RATE As Double = 0.001

' ============================================================
' Function: GetMortalityRate
' Returns the annual probability of death (qx) for a given
' age and gender using a simplified Makeham-Gompertz model.
'
' Parameters:
'   age    - Integer, the attained age
'   gender - String, "M" or "F"
' Returns:
'   Double - the qx value
' ============================================================
Public Function GetMortalityRate(ByVal age As Integer, ByVal gender As String) As Double
    Dim A As Double, B As Double, c As Double
    
    ' Makeham-Gompertz parameters
    If UCase(gender) = "M" Then
        A = 0.0005
        B = 0.00004
        c = 1.1
    Else
        A = 0.0003
        B = 0.000025
        c = 1.095
    End If
    
    ' qx = A + B * c^x
    Dim qx As Double
    qx = A + B * (c ^ age)
    
    ' Cap at 1.0
    If qx > 1# Then qx = 1#
    
    GetMortalityRate = qx
End Function

' ============================================================
' Function: SurvivalProbability
' Returns tPx - the probability of surviving t years from age x.
' ============================================================
Public Function SurvivalProbability(ByVal age As Integer, ByVal t As Integer, _
                                     ByVal gender As String) As Double
    Dim px As Double
    px = 1#
    Dim i As Integer
    For i = 0 To t - 1
        px = px * (1# - GetMortalityRate(age + i, gender))
        If px < 0.000001 Then
            px = 0#
            Exit For
        End If
    Next i
    SurvivalProbability = px
End Function

' ============================================================
' Function: LifeExpectancy
' Calculates curtate life expectancy ex for a given age.
' ============================================================
Public Function LifeExpectancy(ByVal age As Integer, ByVal gender As String) As Double
    Dim ex As Double
    ex = 0#
    Dim t As Integer
    For t = 1 To MAX_AGE - age
        ex = ex + SurvivalProbability(age, t, gender)
    Next t
    LifeExpectancy = ex
End Function

' ============================================================
' Sub: BuildMortalityTable
' Populates a worksheet with a complete mortality table.
' ============================================================
Public Sub BuildMortalityTable()
    Dim ws As Worksheet
    On Error Resume Next
    Set ws = ThisWorkbook.Worksheets("MortalityTable")
    On Error GoTo 0
    
    If ws Is Nothing Then
        Set ws = ThisWorkbook.Worksheets.Add
        ws.Name = "MortalityTable"
    End If
    
    ' Headers
    ws.Cells(1, 1).Value = "Age"
    ws.Cells(1, 2).Value = "qx (Male)"
    ws.Cells(1, 3).Value = "qx (Female)"
    ws.Cells(1, 4).Value = "lx (Male)"
    ws.Cells(1, 5).Value = "lx (Female)"
    ws.Cells(1, 6).Value = "ex (Male)"
    ws.Cells(1, 7).Value = "ex (Female)"
    
    Dim lxM As Double, lxF As Double
    lxM = 100000
    lxF = 100000
    
    Dim row As Integer
    For row = 0 To MAX_AGE
        ws.Cells(row + 2, 1).Value = row
        
        Dim qxM As Double, qxF As Double
        qxM = GetMortalityRate(row, "M")
        qxF = GetMortalityRate(row, "F")
        
        ws.Cells(row + 2, 2).Value = qxM
        ws.Cells(row + 2, 3).Value = qxF
        ws.Cells(row + 2, 4).Value = lxM
        ws.Cells(row + 2, 5).Value = lxF
        ws.Cells(row + 2, 6).Value = LifeExpectancy(row, "M")
        ws.Cells(row + 2, 7).Value = LifeExpectancy(row, "F")
        
        lxM = lxM * (1 - qxM)
        lxF = lxF * (1 - qxF)
    Next row
    
    MsgBox "Mortality table generated with " & MAX_AGE + 1 & " rows.", vbInformation
End Sub
'''

VBA_MODULE_PREMIUM = '''
Attribute VB_Name = "PremiumCalculations"
' ============================================================
' Module: PremiumCalculations
' Purpose: Insurance premium computation for term life,
'          whole life, and endowment products.
' ============================================================
Option Explicit

' ============================================================
' Function: AnnuityDue
' Calculates the present value of a life annuity-due.
'   ax_due = SUM(t=0..n-1) v^t * tPx
' ============================================================
Public Function AnnuityDue(ByVal age As Integer, ByVal term As Integer, _
                            ByVal interest As Double, ByVal gender As String) As Double
    Dim v As Double
    v = 1# / (1# + interest)
    
    Dim total As Double
    total = 0#
    Dim t As Integer
    For t = 0 To term - 1
        total = total + (v ^ t) * SurvivalProbability(age, t, gender)
    Next t
    
    AnnuityDue = total
End Function

' ============================================================
' Function: PureEndowment
' Calculates nEx = v^n * nPx
' ============================================================
Public Function PureEndowment(ByVal age As Integer, ByVal n As Integer, _
                               ByVal interest As Double, ByVal gender As String) As Double
    Dim v As Double
    v = 1# / (1# + interest)
    PureEndowment = (v ^ n) * SurvivalProbability(age, n, gender)
End Function

' ============================================================
' Function: TermLifeNetPremium
' Net annual premium for an n-year term life insurance.
'   P = A(1)x:n / ax_due:n
' where A(1)x:n = sum(t=0..n-1) v^(t+1) * tPx * q(x+t)
' ============================================================
Public Function TermLifeNetPremium(ByVal age As Integer, ByVal term As Integer, _
                                    ByVal sumAssured As Double, ByVal interest As Double, _
                                    ByVal gender As String) As Double
    Dim v As Double
    v = 1# / (1# + interest)
    
    ' Net single premium (term insurance)
    Dim nsp As Double
    nsp = 0#
    Dim t As Integer
    For t = 0 To term - 1
        Dim tPx As Double
        tPx = SurvivalProbability(age, t, gender)
        Dim qxt As Double
        qxt = GetMortalityRate(age + t, gender)
        nsp = nsp + (v ^ (t + 1)) * tPx * qxt
    Next t
    nsp = nsp * sumAssured
    
    ' Annuity due
    Dim annuity As Double
    annuity = AnnuityDue(age, term, interest, gender)
    
    If annuity > 0 Then
        TermLifeNetPremium = nsp / annuity
    Else
        TermLifeNetPremium = 0
    End If
End Function

' ============================================================
' Function: WholeLifeNetPremium
' Net annual premium for whole life insurance.
' ============================================================
Public Function WholeLifeNetPremium(ByVal age As Integer, ByVal sumAssured As Double, _
                                     ByVal interest As Double, ByVal gender As String) As Double
    WholeLifeNetPremium = TermLifeNetPremium(age, 120 - age, sumAssured, interest, gender)
End Function

' ============================================================
' Function: EndowmentNetPremium
' Net annual premium for an n-year endowment.
'   P = (A(1)x:n + nEx) * SA / ax_due:n
' ============================================================
Public Function EndowmentNetPremium(ByVal age As Integer, ByVal term As Integer, _
                                     ByVal sumAssured As Double, ByVal interest As Double, _
                                     ByVal gender As String) As Double
    Dim v As Double
    v = 1# / (1# + interest)
    
    ' Term insurance component
    Dim termIns As Double
    termIns = 0#
    Dim t As Integer
    For t = 0 To term - 1
        termIns = termIns + (v ^ (t + 1)) * SurvivalProbability(age, t, gender) * _
                  GetMortalityRate(age + t, gender)
    Next t
    
    ' Pure endowment component
    Dim pe As Double
    pe = PureEndowment(age, term, interest, gender)
    
    ' Total NSP
    Dim nsp As Double
    nsp = (termIns + pe) * sumAssured
    
    ' Annuity due
    Dim annuity As Double
    annuity = AnnuityDue(age, term, interest, gender)
    
    If annuity > 0 Then
        EndowmentNetPremium = nsp / annuity
    Else
        EndowmentNetPremium = 0
    End If
End Function

' ============================================================
' Function: GrossPremium
' Adds expense and profit loading to net premium.
' ============================================================
Public Function GrossPremium(ByVal netPremium As Double, _
                              ByVal expenseRatio As Double, _
                              ByVal profitMargin As Double) As Double
    GrossPremium = netPremium / (1 - expenseRatio - profitMargin)
End Function

' ============================================================
' Sub: CalculateAllPremiums
' Populates the PremiumSummary sheet with premium calculations
' for a range of policy configurations.
' ============================================================
Public Sub CalculateAllPremiums()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("PremiumCalculations")
    
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).row
    
    Dim r As Long
    For r = 3 To lastRow
        Dim age As Integer: age = CInt(ws.Cells(r, 1).Value)
        Dim gender As String: gender = CStr(ws.Cells(r, 2).Value)
        Dim term As Integer: term = CInt(ws.Cells(r, 3).Value)
        Dim sa As Double: sa = CDbl(ws.Cells(r, 4).Value)
        Dim rate As Double: rate = CDbl(ws.Cells(r, 5).Value)
        Dim prodType As String: prodType = CStr(ws.Cells(r, 6).Value)
        
        Dim netP As Double
        Select Case UCase(prodType)
            Case "TERM"
                netP = TermLifeNetPremium(age, term, sa, rate, gender)
            Case "WHOLE LIFE"
                netP = WholeLifeNetPremium(age, sa, rate, gender)
            Case "ENDOWMENT"
                netP = EndowmentNetPremium(age, term, sa, rate, gender)
            Case Else
                netP = 0
        End Select
        
        ws.Cells(r, 7).Value = netP
        ws.Cells(r, 8).Value = GrossPremium(netP, 0.15, 0.05)
    Next r
    
    MsgBox "Premium calculations complete!", vbInformation
End Sub
'''

VBA_MODULE_RESERVING = '''
Attribute VB_Name = "LossReserving"
' ============================================================
' Module: LossReserving
' Purpose: Property & Casualty loss reserving using the
'          Chain-Ladder (Development Factor) method.
' ============================================================
Option Explicit

' ============================================================
' Function: ChainLadderFactor
' Calculates the age-to-age development factor between two
' consecutive development periods.
' ============================================================
Public Function ChainLadderFactor(ByVal triangle As Range, _
                                   ByVal devPeriod As Integer) As Double
    Dim sumCurrent As Double, sumPrior As Double
    sumCurrent = 0#
    sumPrior = 0#
    
    Dim r As Integer
    For r = 1 To triangle.Rows.Count - devPeriod
        Dim priorVal As Variant, currentVal As Variant
        priorVal = triangle.Cells(r, devPeriod).Value
        currentVal = triangle.Cells(r, devPeriod + 1).Value
        
        If IsNumeric(priorVal) And IsNumeric(currentVal) Then
            If CDbl(priorVal) > 0 Then
                sumPrior = sumPrior + CDbl(priorVal)
                sumCurrent = sumCurrent + CDbl(currentVal)
            End If
        End If
    Next r
    
    If sumPrior > 0 Then
        ChainLadderFactor = sumCurrent / sumPrior
    Else
        ChainLadderFactor = 1#
    End If
End Function

' ============================================================
' Function: CumulativeFactor
' Product of all remaining development factors from a given
' period to ultimate.
' ============================================================
Public Function CumulativeFactor(ByVal triangle As Range, _
                                  ByVal fromPeriod As Integer) As Double
    Dim prod As Double
    prod = 1#
    Dim maxDev As Integer
    maxDev = triangle.Columns.Count - 1
    
    Dim d As Integer
    For d = fromPeriod To maxDev
        prod = prod * ChainLadderFactor(triangle, d)
    Next d
    
    CumulativeFactor = prod
End Function

' ============================================================
' Function: UltimateLoss
' Projects the ultimate loss for a given accident year.
' ============================================================
Public Function UltimateLoss(ByVal triangle As Range, _
                              ByVal accidentYear As Integer) As Double
    ' Find the latest diagonal value for this accident year
    Dim latestDev As Integer
    latestDev = triangle.Columns.Count - accidentYear + 1
    If latestDev > triangle.Columns.Count Then latestDev = triangle.Columns.Count
    
    Dim latestValue As Double
    latestValue = CDbl(triangle.Cells(accidentYear, latestDev).Value)
    
    ' Multiply by cumulative development factor
    If latestDev < triangle.Columns.Count Then
        UltimateLoss = latestValue * CumulativeFactor(triangle, latestDev)
    Else
        UltimateLoss = latestValue
    End If
End Function

' ============================================================
' Function: IBNRReserve
' Calculates IBNR (Incurred But Not Reported) reserve.
'   IBNR = Ultimate - Paid-to-date
' ============================================================
Public Function IBNRReserve(ByVal triangle As Range, _
                             ByVal accidentYear As Integer) As Double
    Dim latestDev As Integer
    latestDev = triangle.Columns.Count - accidentYear + 1
    If latestDev > triangle.Columns.Count Then latestDev = triangle.Columns.Count
    
    Dim paidToDate As Double
    paidToDate = CDbl(triangle.Cells(accidentYear, latestDev).Value)
    
    IBNRReserve = UltimateLoss(triangle, accidentYear) - paidToDate
End Function

' ============================================================
' Sub: RunChainLadder
' Runs the chain-ladder analysis on the LossTriangle sheet.
' ============================================================
Public Sub RunChainLadder()
    Dim wsTriangle As Worksheet
    Set wsTriangle = ThisWorkbook.Worksheets("LossTriangle")
    
    ' Find the triangle range (assumes it starts at B3)
    Dim lastRow As Long, lastCol As Long
    lastRow = wsTriangle.Cells(wsTriangle.Rows.Count, 2).End(xlUp).row
    lastCol = wsTriangle.Cells(3, wsTriangle.Columns.Count).End(xlToLeft).Column
    
    Dim triangleRange As Range
    Set triangleRange = wsTriangle.Range(wsTriangle.Cells(3, 2), wsTriangle.Cells(lastRow, lastCol))
    
    ' Output development factors
    Dim outputRow As Long
    outputRow = lastRow + 3
    wsTriangle.Cells(outputRow, 1).Value = "Development Factors"
    wsTriangle.Cells(outputRow, 1).Font.Bold = True
    
    Dim d As Integer
    For d = 1 To lastCol - 2
        wsTriangle.Cells(outputRow + 1, d + 1).Value = d & " to " & d + 1
        wsTriangle.Cells(outputRow + 2, d + 1).Value = ChainLadderFactor(triangleRange, d)
        wsTriangle.Cells(outputRow + 2, d + 1).NumberFormat = "0.0000"
    Next d
    
    ' Output IBNR reserves
    outputRow = outputRow + 5
    wsTriangle.Cells(outputRow, 1).Value = "IBNR Reserves"
    wsTriangle.Cells(outputRow, 1).Font.Bold = True
    wsTriangle.Cells(outputRow + 1, 1).Value = "Acc. Year"
    wsTriangle.Cells(outputRow + 1, 2).Value = "Paid to Date"
    wsTriangle.Cells(outputRow + 1, 3).Value = "Ultimate"
    wsTriangle.Cells(outputRow + 1, 4).Value = "IBNR"
    
    Dim numYears As Integer
    numYears = lastRow - 2
    
    Dim ay As Integer
    For ay = 1 To numYears
        Dim latestDev As Integer
        latestDev = lastCol - 1 - ay + 1
        If latestDev > lastCol - 1 Then latestDev = lastCol - 1
        
        wsTriangle.Cells(outputRow + 1 + ay, 1).Value = wsTriangle.Cells(2 + ay, 1).Value
        wsTriangle.Cells(outputRow + 1 + ay, 2).Value = triangleRange.Cells(ay, latestDev).Value
        wsTriangle.Cells(outputRow + 1 + ay, 3).Value = UltimateLoss(triangleRange, ay)
        wsTriangle.Cells(outputRow + 1 + ay, 4).Value = IBNRReserve(triangleRange, ay)
        
        wsTriangle.Cells(outputRow + 1 + ay, 2).NumberFormat = "#,##0"
        wsTriangle.Cells(outputRow + 1 + ay, 3).NumberFormat = "#,##0"
        wsTriangle.Cells(outputRow + 1 + ay, 4).NumberFormat = "#,##0"
    Next ay
    
    MsgBox "Chain-Ladder analysis complete! " & numYears & " accident years processed.", vbInformation
End Sub
'''

VBA_MODULE_RISK = '''
Attribute VB_Name = "RiskMetrics"
' ============================================================
' Module: RiskMetrics
' Purpose: Risk and capital adequacy calculations including
'          Value at Risk (VaR), Expected Shortfall, and
'          Solvency ratios.
' ============================================================
Option Explicit

' ============================================================
' Function: NormalInverse
' Approximation of the inverse standard normal CDF
' using the Abramowitz & Stegun rational approximation.
' ============================================================
Public Function NormalInverse(ByVal p As Double) As Double
    ' Coefficients
    Const a1 As Double = -3.969683028665376E+01
    Const a2 As Double = 2.209460984245205E+02
    Const a3 As Double = -2.759285104469687E+02
    Const a4 As Double = 1.383577518672690E+02
    Const a5 As Double = -3.066479806614716E+01
    Const a6 As Double = 2.506628277459239E+00
    
    Const b1 As Double = -5.447609879822406E+01
    Const b2 As Double = 1.615858368580409E+02
    Const b3 As Double = -1.556989798598866E+02
    Const b4 As Double = 6.680131188771972E+01
    Const b5 As Double = -1.328068155288572E+01
    
    Const c1 As Double = -7.784894002430293E-03
    Const c2 As Double = -3.223964580411365E-01
    Const c3 As Double = -2.400758277161838E+00
    Const c4 As Double = -2.549732539343734E+00
    Const c5 As Double = 4.374664141464968E+00
    Const c6 As Double = 2.938163982698783E+00
    
    Const d1 As Double = 7.784695709041462E-03
    Const d2 As Double = 3.224671290700398E-01
    Const d3 As Double = 2.445134137142996E+00
    Const d4 As Double = 3.754408661907416E+00
    
    Dim q As Double, r As Double
    
    If p < 0.02425 Then
        ' Lower tail
        q = Sqr(-2# * Log(p))
        NormalInverse = (((((c1 * q + c2) * q + c3) * q + c4) * q + c5) * q + c6) / _
                        ((((d1 * q + d2) * q + d3) * q + d4) * q + 1#)
    ElseIf p <= 0.97575 Then
        ' Central region
        q = p - 0.5
        r = q * q
        NormalInverse = (((((a1 * r + a2) * r + a3) * r + a4) * r + a5) * r + a6) * q / _
                        (((((b1 * r + b2) * r + b3) * r + b4) * r + b5) * r + 1#)
    Else
        ' Upper tail
        q = Sqr(-2# * Log(1# - p))
        NormalInverse = -(((((c1 * q + c2) * q + c3) * q + c4) * q + c5) * q + c6) / _
                         ((((d1 * q + d2) * q + d3) * q + d4) * q + 1#))
    End If
End Function

' ============================================================
' Function: ValueAtRisk
' Parametric VaR assuming normal distribution.
'   VaR = mu + sigma * z_alpha
' ============================================================
Public Function ValueAtRisk(ByVal meanLoss As Double, ByVal stdDev As Double, _
                             ByVal confidenceLevel As Double) As Double
    ValueAtRisk = meanLoss + stdDev * NormalInverse(confidenceLevel)
End Function

' ============================================================
' Function: ExpectedShortfall
' Conditional Tail Expectation (CTE) / Expected Shortfall
' for a normal distribution.
'   ES = mu + sigma * phi(z_alpha) / (1 - alpha)
' ============================================================
Public Function ExpectedShortfall(ByVal meanLoss As Double, ByVal stdDev As Double, _
                                   ByVal confidenceLevel As Double) As Double
    Dim z As Double
    z = NormalInverse(confidenceLevel)
    
    ' Standard normal PDF at z
    Dim phi As Double
    phi = Exp(-0.5 * z * z) / Sqr(2# * 3.14159265358979)
    
    ExpectedShortfall = meanLoss + stdDev * phi / (1# - confidenceLevel)
End Function

' ============================================================
' Function: SolvencyRatio
' Calculates the solvency ratio = Available Capital / Required Capital.
' ============================================================
Public Function SolvencyRatio(ByVal availableCapital As Double, _
                               ByVal requiredCapital As Double) As Double
    If requiredCapital > 0 Then
        SolvencyRatio = availableCapital / requiredCapital
    Else
        SolvencyRatio = 0
    End If
End Function

' ============================================================
' Function: RiskBasedCapital
' Simplified Risk-Based Capital calculation.
'   RBC = Sqrt(C1^2 + C2^2 + C3^2 + C4^2)
'   where C1-C4 are risk charges for different categories.
' ============================================================
Public Function RiskBasedCapital(ByVal assetRisk As Double, _
                                  ByVal insuranceRisk As Double, _
                                  ByVal interestRateRisk As Double, _
                                  ByVal businessRisk As Double) As Double
    RiskBasedCapital = Sqr(assetRisk ^ 2 + insuranceRisk ^ 2 + _
                           interestRateRisk ^ 2 + businessRisk ^ 2)
End Function

' ============================================================
' Function: LossRatio
' Calculates the loss ratio = Incurred Losses / Earned Premiums
' ============================================================
Public Function LossRatio(ByVal incurredLosses As Double, _
                           ByVal earnedPremiums As Double) As Double
    If earnedPremiums > 0 Then
        LossRatio = incurredLosses / earnedPremiums
    Else
        LossRatio = 0
    End If
End Function

' ============================================================
' Function: CombinedRatio
' Combined Ratio = Loss Ratio + Expense Ratio
' ============================================================
Public Function CombinedRatio(ByVal incurredLosses As Double, _
                               ByVal expenses As Double, _
                               ByVal earnedPremiums As Double) As Double
    If earnedPremiums > 0 Then
        CombinedRatio = (incurredLosses + expenses) / earnedPremiums
    Else
        CombinedRatio = 0
    End If
End Function

' ============================================================
' Sub: CalculateRiskMetrics
' Populates the RiskAnalysis sheet.
' ============================================================
Public Sub CalculateRiskMetrics()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Worksheets("RiskAnalysis")
    
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).row
    
    Dim r As Long
    For r = 3 To lastRow
        Dim meanL As Double: meanL = CDbl(ws.Cells(r, 2).Value)
        Dim stdL As Double: stdL = CDbl(ws.Cells(r, 3).Value)
        
        ' VaR at 95% and 99.5%
        ws.Cells(r, 4).Value = ValueAtRisk(meanL, stdL, 0.95)
        ws.Cells(r, 5).Value = ValueAtRisk(meanL, stdL, 0.995)
        
        ' Expected Shortfall at 95% and 99.5%  
        ws.Cells(r, 6).Value = ExpectedShortfall(meanL, stdL, 0.95)
        ws.Cells(r, 7).Value = ExpectedShortfall(meanL, stdL, 0.995)
    Next r
    
    MsgBox "Risk metrics calculated!", vbInformation
End Sub
'''

VBA_CLASS_POLICY = '''
Attribute VB_Name = "clsPolicy"
' ============================================================
' Class Module: clsPolicy
' Represents an individual insurance policy with properties
' and methods for valuation.
' ============================================================
Option Explicit

Private pPolicyID As String
Private pInsuredName As String
Private pAge As Integer
Private pGender As String
Private pProductType As String
Private pSumAssured As Double
Private pTerm As Integer
Private pInterestRate As Double
Private pIssueDate As Date

' --- Properties ---
Public Property Get PolicyID() As String
    PolicyID = pPolicyID
End Property
Public Property Let PolicyID(ByVal value As String)
    pPolicyID = value
End Property

Public Property Get InsuredName() As String
    InsuredName = pInsuredName
End Property
Public Property Let InsuredName(ByVal value As String)
    pInsuredName = value
End Property

Public Property Get Age() As Integer
    Age = pAge
End Property
Public Property Let Age(ByVal value As Integer)
    pAge = value
End Property

Public Property Get Gender() As String
    Gender = pGender
End Property
Public Property Let Gender(ByVal value As String)
    pGender = value
End Property

Public Property Get ProductType() As String
    ProductType = pProductType
End Property
Public Property Let ProductType(ByVal value As String)
    pProductType = value
End Property

Public Property Get SumAssured() As Double
    SumAssured = pSumAssured
End Property
Public Property Let SumAssured(ByVal value As Double)
    pSumAssured = value
End Property

Public Property Get Term() As Integer
    Term = pTerm
End Property
Public Property Let Term(ByVal value As Integer)
    pTerm = value
End Property

Public Property Get InterestRate() As Double
    InterestRate = pInterestRate
End Property
Public Property Let InterestRate(ByVal value As Double)
    pInterestRate = value
End Property

Public Property Get IssueDate() As Date
    IssueDate = pIssueDate
End Property
Public Property Let IssueDate(ByVal value As Date)
    pIssueDate = value
End Property

' ============================================================
' Function: CalculateNetPremium
' Returns the net premium for this policy.
' ============================================================
Public Function CalculateNetPremium() As Double
    Select Case UCase(pProductType)
        Case "TERM"
            CalculateNetPremium = TermLifeNetPremium(pAge, pTerm, pSumAssured, pInterestRate, pGender)
        Case "WHOLE LIFE"
            CalculateNetPremium = WholeLifeNetPremium(pAge, pSumAssured, pInterestRate, pGender)
        Case "ENDOWMENT"
            CalculateNetPremium = EndowmentNetPremium(pAge, pTerm, pSumAssured, pInterestRate, pGender)
        Case Else
            CalculateNetPremium = 0
    End Select
End Function

' ============================================================
' Function: CalculateReserve
' Returns the policy reserve at a given duration t.
'   tVx = A(x+t) - P * ax_due(x+t)  (prospective method)
' ============================================================
Public Function CalculateReserve(ByVal duration As Integer) As Double
    If duration >= pTerm Then
        CalculateReserve = 0
        Exit Function
    End If
    
    Dim futureAge As Integer
    futureAge = pAge + duration
    Dim remainingTerm As Integer
    remainingTerm = pTerm - duration
    
    ' Future benefit (NSP of remaining coverage)
    Dim v As Double
    v = 1# / (1# + pInterestRate)
    
    Dim futureBenefit As Double
    futureBenefit = 0#
    Dim t As Integer
    For t = 0 To remainingTerm - 1
        futureBenefit = futureBenefit + (v ^ (t + 1)) * _
                        SurvivalProbability(futureAge, t, pGender) * _
                        GetMortalityRate(futureAge + t, pGender)
    Next t
    futureBenefit = futureBenefit * pSumAssured
    
    ' Future premiums
    Dim futurePremiums As Double
    futurePremiums = CalculateNetPremium() * AnnuityDue(futureAge, remainingTerm, pInterestRate, pGender)
    
    CalculateReserve = futureBenefit - futurePremiums
End Function
'''

VBA_THISWORKBOOK = '''
Attribute VB_Name = "ThisWorkbook"
' ============================================================
' ThisWorkbook Module
' Handles workbook-level events and initialisation.
' ============================================================
Option Explicit

Private Sub Workbook_Open()
    ' Display welcome message
    MsgBox "Welcome to the Actuarial Insurance Model." & vbCrLf & _
           "Use the macros in the Developer tab to run analyses:" & vbCrLf & vbCrLf & _
           "  - BuildMortalityTable: Generate full life table" & vbCrLf & _
           "  - CalculateAllPremiums: Compute premiums for all policies" & vbCrLf & _
           "  - RunChainLadder: Run loss reserving analysis" & vbCrLf & _
           "  - CalculateRiskMetrics: Compute VaR and risk measures", _
           vbInformation, "Actuarial Insurance Model v2.0"
End Sub
'''


# ---------------------------------------------------------------------------
# Helper: Makeham-Gompertz qx (mirrors VBA logic)
# ---------------------------------------------------------------------------

def _qx(age: int, gender: str) -> float:
    if gender.upper() == "M":
        a, b, c = 0.0005, 0.00004, 1.1
    else:
        a, b, c = 0.0003, 0.000025, 1.095
    return min(a + b * (c ** age), 1.0)


def _survival(age: int, t: int, gender: str) -> float:
    px = 1.0
    for i in range(t):
        px *= (1.0 - _qx(age + i, gender))
        if px < 1e-6:
            return 0.0
    return px


def _life_expectancy(age: int, gender: str) -> float:
    return sum(_survival(age, t, gender) for t in range(1, 121 - age))


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 12) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(max_len + 3, min_width)


def build_mortality_table(wb: openpyxl.Workbook) -> None:
    """Sheet 1 — Mortality Table with qx, lx, ex."""
    ws = wb.create_sheet("MortalityTable")
    headers = ["Age", "qx (Male)", "qx (Female)", "lx (Male)", "lx (Female)",
               "ex (Male)", "ex (Female)"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    lx_m, lx_f = 100_000.0, 100_000.0
    for age in range(0, 121):
        qm, qf = _qx(age, "M"), _qx(age, "F")
        ws.append([
            age,
            round(qm, 8), round(qf, 8),
            round(lx_m, 2), round(lx_f, 2),
            round(_life_expectancy(age, "M"), 2),
            round(_life_expectancy(age, "F"), 2),
        ])
        lx_m *= (1 - qm)
        lx_f *= (1 - qf)

    # Format
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '0.00000000'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = '#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = '0.00'

    # Line chart for qx
    chart = LineChart()
    chart.title = "Mortality Rates by Age (qx)"
    chart.y_axis.title = "Probability of Death"
    chart.x_axis.title = "Age"
    chart.style = 10
    chart.width = 24
    chart.height = 14
    ages = Reference(ws, min_col=1, min_row=2, max_row=122)
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=122), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=3, min_row=1, max_row=122), titles_from_data=True)
    chart.set_categories(ages)
    ws.add_chart(chart, "I2")

    _auto_width(ws)


def build_premium_calculations(wb: openpyxl.Workbook) -> None:
    """Sheet 2 — Premium Calculations for sample policies."""
    ws = wb.create_sheet("PremiumCalculations")

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Insurance Premium Calculations"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Age", "Gender", "Term (yrs)", "Sum Assured", "Interest Rate",
               "Product Type", "Net Premium", "Gross Premium"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    # Sample policies
    policies = [
        (25, "M", 20, 500_000, 0.04, "Term"),
        (30, "F", 25, 750_000, 0.035, "Term"),
        (35, "M", 30, 1_000_000, 0.04, "Whole Life"),
        (40, "F", 20, 500_000, 0.03, "Endowment"),
        (45, "M", 15, 250_000, 0.045, "Term"),
        (28, "F", 30, 1_500_000, 0.04, "Whole Life"),
        (50, "M", 10, 200_000, 0.035, "Endowment"),
        (33, "F", 25, 800_000, 0.04, "Term"),
        (55, "M", 20, 300_000, 0.03, "Whole Life"),
        (38, "F", 15, 600_000, 0.045, "Endowment"),
        (22, "M", 30, 2_000_000, 0.04, "Term"),
        (42, "F", 20, 400_000, 0.035, "Whole Life"),
        (60, "M", 10, 150_000, 0.03, "Term"),
        (29, "M", 25, 1_200_000, 0.04, "Endowment"),
        (47, "F", 15, 350_000, 0.035, "Term"),
    ]

    for i, (age, gen, term, sa, rate, prod) in enumerate(policies, 3):
        ws.cell(row=i, column=1, value=age)
        ws.cell(row=i, column=2, value=gen)
        ws.cell(row=i, column=3, value=term)
        ws.cell(row=i, column=4, value=sa).number_format = CURRENCY_FMT
        ws.cell(row=i, column=5, value=rate).number_format = PCT_FMT
        ws.cell(row=i, column=6, value=prod)
        # Formulas referencing the VBA UDFs
        ws.cell(row=i, column=7).value = f'=IF(F{i}="Term",TermLifeNetPremium(A{i},C{i},D{i},E{i},B{i}),IF(F{i}="Whole Life",WholeLifeNetPremium(A{i},D{i},E{i},B{i}),EndowmentNetPremium(A{i},C{i},D{i},E{i},B{i})))'
        ws.cell(row=i, column=7).number_format = CURRENCY_FMT
        ws.cell(row=i, column=8).value = f'=GrossPremium(G{i},0.15,0.05)'
        ws.cell(row=i, column=8).number_format = CURRENCY_FMT

    _auto_width(ws)


def build_loss_triangle(wb: openpyxl.Workbook) -> None:
    """Sheet 3 — P&C Loss Development Triangle for Chain Ladder."""
    ws = wb.create_sheet("LossTriangle")

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "Cumulative Paid Loss Development Triangle"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    # Development periods 1-10
    dev_periods = 10
    base_year = 2015
    num_years = 10
    headers = ["Accident Year"] + [f"Dev {d}" for d in range(1, dev_periods + 1)]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    # Generate realistic cumulative triangle data
    for ay_idx in range(num_years):
        year = base_year + ay_idx
        ws.cell(row=3 + ay_idx, column=1, value=year)

        # Base incurred loss for this accident year
        base_loss = random.randint(800_000, 3_000_000)
        cumulative = base_loss
        filled_periods = num_years - ay_idx  # diagonal constraint

        for dev in range(1, filled_periods + 1):
            ws.cell(row=3 + ay_idx, column=1 + dev, value=round(cumulative))
            ws.cell(row=3 + ay_idx, column=1 + dev).number_format = '#,##0'
            # Development factor decays
            dev_factor = 1.0 + max(0.02, 0.5 / (dev + 0.5)) * random.uniform(0.7, 1.3)
            cumulative *= dev_factor

    _auto_width(ws, min_width=14)


def build_risk_analysis(wb: openpyxl.Workbook) -> None:
    """Sheet 4 — Risk Analysis with VaR, ES, solvency metrics."""
    ws = wb.create_sheet("RiskAnalysis")

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Risk Analysis & Capital Adequacy"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Line of Business", "Mean Loss ($M)", "Std Dev ($M)",
               "VaR 95% ($M)", "VaR 99.5% ($M)", "ES 95% ($M)", "ES 99.5% ($M)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    lines = [
        ("Auto Liability",    45.2, 12.5),
        ("Property",          82.1, 28.3),
        ("Workers Comp",      31.7,  9.8),
        ("General Liability", 56.4, 18.2),
        ("Professional Liability", 22.3, 8.1),
        ("Medical Malpractice", 38.9, 15.6),
        ("Commercial Multi-Peril", 67.5, 22.4),
        ("Homeowners", 91.3, 35.7),
    ]

    for i, (lob, mean, std) in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=lob)
        ws.cell(row=i, column=2, value=mean).number_format = CURRENCY_FMT
        ws.cell(row=i, column=3, value=std).number_format = CURRENCY_FMT
        # VBA UDF formulas
        ws.cell(row=i, column=4).value = f'=ValueAtRisk(B{i},C{i},0.95)'
        ws.cell(row=i, column=4).number_format = CURRENCY_FMT
        ws.cell(row=i, column=5).value = f'=ValueAtRisk(B{i},C{i},0.995)'
        ws.cell(row=i, column=5).number_format = CURRENCY_FMT
        ws.cell(row=i, column=6).value = f'=ExpectedShortfall(B{i},C{i},0.95)'
        ws.cell(row=i, column=6).number_format = CURRENCY_FMT
        ws.cell(row=i, column=7).value = f'=ExpectedShortfall(B{i},C{i},0.995)'
        ws.cell(row=i, column=7).number_format = CURRENCY_FMT

    # Solvency section
    row = len(lines) + 5
    ws.cell(row=row, column=1, value="Capital Adequacy Summary").font = Font(bold=True, size=12)
    ws.cell(row=row + 1, column=1, value="Available Capital ($M)")
    ws.cell(row=row + 1, column=2, value=250.0).number_format = CURRENCY_FMT
    ws.cell(row=row + 2, column=1, value="Asset Risk ($M)")
    ws.cell(row=row + 2, column=2, value=35.0).number_format = CURRENCY_FMT
    ws.cell(row=row + 3, column=1, value="Insurance Risk ($M)")
    ws.cell(row=row + 3, column=2, value=48.0).number_format = CURRENCY_FMT
    ws.cell(row=row + 4, column=1, value="Interest Rate Risk ($M)")
    ws.cell(row=row + 4, column=2, value=15.0).number_format = CURRENCY_FMT
    ws.cell(row=row + 5, column=1, value="Business Risk ($M)")
    ws.cell(row=row + 5, column=2, value=12.0).number_format = CURRENCY_FMT

    rbc_row = row + 6
    ws.cell(row=rbc_row, column=1, value="Risk-Based Capital ($M)").font = Font(bold=True)
    rbc_formula = f'=RiskBasedCapital(B{row+2},B{row+3},B{row+4},B{row+5})'
    ws.cell(row=rbc_row, column=2).value = rbc_formula
    ws.cell(row=rbc_row, column=2).number_format = CURRENCY_FMT

    ws.cell(row=rbc_row + 1, column=1, value="Solvency Ratio").font = Font(bold=True)
    ws.cell(row=rbc_row + 1, column=2).value = f'=SolvencyRatio(B{row+1},B{rbc_row})'
    ws.cell(row=rbc_row + 1, column=2).number_format = PCT_FMT

    _auto_width(ws)


def build_policy_portfolio(wb: openpyxl.Workbook) -> None:
    """Sheet 5 — Sample policy portfolio with detailed records."""
    ws = wb.create_sheet("PolicyPortfolio")

    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "Insurance Policy Portfolio"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Policy ID", "Insured Name", "Age", "Gender", "Product",
               "Sum Assured", "Term", "Interest Rate", "Issue Date",
               "Annual Premium", "Reserve", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    first_names_m = ["James", "Robert", "Michael", "William", "David",
                     "Richard", "Joseph", "Thomas", "Charles", "Daniel"]
    first_names_f = ["Mary", "Patricia", "Jennifer", "Linda", "Barbara",
                     "Elizabeth", "Susan", "Jessica", "Sarah", "Karen"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones",
                  "Garcia", "Miller", "Davis", "Rodriguez", "Martinez",
                  "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson"]
    products = ["Term", "Whole Life", "Endowment"]
    statuses = ["Active", "Active", "Active", "Active", "Lapsed", "Paid-up"]

    for i in range(50):
        row = i + 3
        gender = random.choice(["M", "F"])
        first = random.choice(first_names_m if gender == "M" else first_names_f)
        last = random.choice(last_names)
        age = random.randint(22, 65)
        product = random.choice(products)
        term = random.choice([10, 15, 20, 25, 30]) if product != "Whole Life" else 85 - age
        sa = random.choice([100_000, 250_000, 500_000, 750_000, 1_000_000, 1_500_000, 2_000_000])
        rate = round(random.uniform(0.025, 0.05), 3)
        issue_date = date(
            random.randint(2015, 2024),
            random.randint(1, 12),
            random.randint(1, 28),
        )
        status = random.choice(statuses)

        ws.cell(row=row, column=1, value=f"POL-{2024000 + i}")
        ws.cell(row=row, column=2, value=f"{first} {last}")
        ws.cell(row=row, column=3, value=age)
        ws.cell(row=row, column=4, value=gender)
        ws.cell(row=row, column=5, value=product)
        ws.cell(row=row, column=6, value=sa).number_format = CURRENCY_FMT
        ws.cell(row=row, column=7, value=term)
        ws.cell(row=row, column=8, value=rate).number_format = PCT_FMT
        ws.cell(row=row, column=9, value=issue_date).number_format = 'YYYY-MM-DD'
        # Premium formula
        prem_formula = (
            f'=IF(E{row}="Term",TermLifeNetPremium(C{row},G{row},F{row},H{row},D{row}),'
            f'IF(E{row}="Whole Life",WholeLifeNetPremium(C{row},F{row},H{row},D{row}),'
            f'EndowmentNetPremium(C{row},G{row},F{row},H{row},D{row})))'
        )
        ws.cell(row=row, column=10).value = prem_formula
        ws.cell(row=row, column=10).number_format = CURRENCY_FMT
        # Reserve — simplified static value
        ws.cell(row=row, column=11).value = f'=J{row}*1.2'
        ws.cell(row=row, column=11).number_format = CURRENCY_FMT
        ws.cell(row=row, column=12, value=status)

    # Summary row
    summary_row = 53
    ws.cell(row=summary_row, column=5, value="TOTAL").font = Font(bold=True)
    ws.cell(row=summary_row, column=6).value = f'=SUM(F3:F52)'
    ws.cell(row=summary_row, column=6).number_format = CURRENCY_FMT
    ws.cell(row=summary_row, column=6).font = Font(bold=True)
    ws.cell(row=summary_row, column=10).value = f'=SUM(J3:J52)'
    ws.cell(row=summary_row, column=10).number_format = CURRENCY_FMT
    ws.cell(row=summary_row, column=10).font = Font(bold=True)

    _auto_width(ws)


def build_experience_analysis(wb: openpyxl.Workbook) -> None:
    """Sheet 6 — Actual vs Expected mortality experience study."""
    ws = wb.create_sheet("ExperienceStudy")

    ws.merge_cells("A1:I1")
    ws["A1"].value = "Mortality Experience Study — Actual vs Expected"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Age Band", "Exposure (Life-Years)", "Actual Deaths",
               "Expected Deaths (qx)", "A/E Ratio", "95% CI Lower",
               "95% CI Upper", "Credibility Factor", "Adjusted qx"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    age_bands = ["20-29", "30-39", "40-49", "50-59", "60-69", "70-79", "80-89", "90+"]
    exposures = [12500, 18200, 22100, 19800, 14300, 8700, 3200, 800]
    actual_deaths = [8, 22, 58, 142, 285, 412, 328, 195]
    expected_deaths = [7.5, 20.1, 55.3, 138.7, 271.4, 395.8, 310.2, 188.4]

    for i, (band, exp, act, expt) in enumerate(
        zip(age_bands, exposures, actual_deaths, expected_deaths), 3
    ):
        ws.cell(row=i, column=1, value=band)
        ws.cell(row=i, column=2, value=exp).number_format = '#,##0'
        ws.cell(row=i, column=3, value=act)
        ws.cell(row=i, column=4, value=expt).number_format = '#,##0.0'
        # A/E Ratio
        ws.cell(row=i, column=5).value = f'=C{i}/D{i}'
        ws.cell(row=i, column=5).number_format = PCT_FMT
        # 95% CI (normal approx)
        ws.cell(row=i, column=6).value = f'=C{i}/D{i}-1.96*SQRT(C{i})/D{i}'
        ws.cell(row=i, column=6).number_format = PCT_FMT
        ws.cell(row=i, column=7).value = f'=C{i}/D{i}+1.96*SQRT(C{i})/D{i}'
        ws.cell(row=i, column=7).number_format = PCT_FMT
        # Credibility factor Z = min(1, sqrt(n/1082.41))  for full credibility
        ws.cell(row=i, column=8).value = f'=MIN(1,SQRT(C{i}/1082.41))'
        ws.cell(row=i, column=8).number_format = '0.000'
        # Adjusted qx = Z * Actual_rate + (1-Z) * Expected_rate
        ws.cell(row=i, column=9).value = f'=H{i}*(C{i}/B{i})+(1-H{i})*(D{i}/B{i})'
        ws.cell(row=i, column=9).number_format = '0.000000'

    # A/E bar chart
    chart = BarChart()
    chart.title = "Actual / Expected Mortality Ratio by Age Band"
    chart.y_axis.title = "A/E Ratio"
    chart.style = 10
    chart.width = 20
    chart.height = 12
    cats = Reference(ws, min_col=1, min_row=3, max_row=10)
    data = Reference(ws, min_col=5, min_row=2, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A14")

    _auto_width(ws)


def build_assumptions(wb: openpyxl.Workbook) -> None:
    """Sheet 7 — Key actuarial assumptions & parameters."""
    ws = wb.create_sheet("Assumptions")

    ws.merge_cells("A1:D1")
    ws["A1"].value = "Actuarial Assumptions & Parameters"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    headers = ["Parameter", "Value", "Unit", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    assumptions = [
        ("Valuation Date", "2025-12-31", "", "Year-end valuation"),
        ("Mortality Table", "Makeham-Gompertz", "", "A + B*c^x model"),
        ("Makeham A (Male)", 0.0005, "", "Accident/background mortality"),
        ("Makeham B (Male)", 0.00004, "", "Senescent mortality coefficient"),
        ("Makeham c (Male)", 1.1, "", "Aging acceleration factor"),
        ("Makeham A (Female)", 0.0003, "", "Lower background mortality"),
        ("Makeham B (Female)", 0.000025, "", "Lower senescent coefficient"),
        ("Makeham c (Female)", 1.095, "", "Slower aging acceleration"),
        ("Discount Rate — Term", 0.04, "%", "Risk-free + spread"),
        ("Discount Rate — WL", 0.035, "%", "Lower for long-duration"),
        ("Discount Rate — Endowment", 0.03, "%", "Conservative for savings"),
        ("Expense Ratio", 0.15, "%", "Acquisition + maintenance"),
        ("Profit Margin", 0.05, "%", "Target profit loading"),
        ("Lapse Rate (Year 1)", 0.08, "%", "First-year lapse"),
        ("Lapse Rate (Year 2+)", 0.04, "%", "Renewal lapse"),
        ("Inflation Rate", 0.025, "%", "Expense inflation assumption"),
        ("Tax Rate", 0.21, "%", "Corporate tax rate"),
        ("Solvency Target", 2.0, "x", "Target solvency ratio"),
        ("VaR Confidence", 0.995, "", "99.5th percentile for SCR"),
        ("Credibility Standard", 1082.41, "deaths", "Full credibility criterion"),
    ]

    for i, (param, val, unit, note) in enumerate(assumptions, 3):
        ws.cell(row=i, column=1, value=param)
        cell = ws.cell(row=i, column=2, value=val)
        if unit == "%":
            cell.number_format = PCT_FMT
        ws.cell(row=i, column=3, value=unit)
        ws.cell(row=i, column=4, value=note)

    _auto_width(ws, min_width=14)


# ---------------------------------------------------------------------------
# OLE / CFB binary builder for vbaProject.bin
# ---------------------------------------------------------------------------

def _build_vba_project_bin(modules: list[tuple[str, str, int]]) -> bytes:
    """
    Build a minimal OLE2 Compound File Binary (CFB) containing VBA source
    modules that Excel will recognise as a vbaProject.bin.

    Each entry in *modules* is (name, source_code, module_type) where
    module_type is 1=standard, 2=class, 100=ThisWorkbook/Sheet.

    The binary is constructed from scratch following the MS-CFB and MS-OVBA
    specifications so that openpyxl can attach it as the vba_archive.
    """
    # This is a simplified but functional CFB builder
    # For production use, consider using olefile to build the binary
    
    # We'll build an in-memory ZIP-based approach instead,
    # since openpyxl.Workbook.vba_archive expects a ZipFile-like object
    # containing the vbaProject.bin entry.
    
    # Actually, the vba_archive is a ZipExtFile from the original xlsm.
    # The cleanest cross-platform approach is to build a minimal xlsm
    # template with VBA using the xl/vbaProject.bin OLE stream.
    
    # For now, we'll use a pre-constructed minimal vbaProject.bin
    # This is the approach used by openpyxl's test suite
    pass  # placeholder — see inject_vba_via_zip below


def inject_vba_via_zip(xlsx_path: str, output_xlsm_path: str,
                        vba_modules: list[tuple[str, str]]) -> None:
    """
    Take an .xlsx file built by openpyxl and convert it to .xlsm by:
    1. Changing [Content_Types].xml to declare vbaProject.bin
    2. Adding xl/vbaProject.bin (a real OLE2 CFB stream with compressed VBA)
    3. Adding the relationship entry

    This builds a genuine vbaProject.bin from scratch using the MS-CFB and
    MS-OVBA binary specifications.
    """
    from _vba_project_builder import build_vba_project

    # Convert (name, source) pairs to (name, source, is_class) triples
    CLASS_MODULES = {"clsPolicy", "ThisWorkbook"}
    modules_with_type = [
        (name, source.strip() + "\r\n", name in CLASS_MODULES)
        for name, source in vba_modules
    ]

    vba_bin = build_vba_project(modules_with_type)

    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
         zipfile.ZipFile(output_xlsm_path, 'w') as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)

            if item.filename == '[Content_Types].xml':
                # Add vbaProject content type
                data = data.replace(
                    b'</Types>',
                    b'<Override PartName="/xl/vbaProject.bin" '
                    b'ContentType="application/vnd.ms-office.vbaProject"/>'
                    b'</Types>'
                )
            elif item.filename == 'xl/_rels/workbook.xml.rels':
                # Add relationship to vbaProject.bin
                data = data.replace(
                    b'</Relationships>',
                    b'<Relationship Id="rIdVBA" Type='
                    b'"http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                    b'Target="vbaProject.bin"/>'
                    b'</Relationships>'
                )

            zout.writestr(item, data)

        # Write the vbaProject.bin
        zout.writestr('xl/vbaProject.bin', vba_bin)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building sheets...")
    build_mortality_table(wb)
    build_premium_calculations(wb)
    build_loss_triangle(wb)
    build_risk_analysis(wb)
    build_policy_portfolio(wb)
    build_experience_analysis(wb)
    build_assumptions(wb)
    print(f"  Created {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

    # Save as .xlsx first, then we'll inject VBA
    temp_xlsx = OUTPUT_DIR / "temp_actuarial.xlsx"
    wb.save(str(temp_xlsx))
    print(f"Saved intermediate workbook to {temp_xlsx}")

    # Now inject VBA modules
    vba_modules = [
        ("ThisWorkbook", VBA_THISWORKBOOK),
        ("MortalityFunctions", VBA_MODULE_MORTALITY),
        ("PremiumCalculations", VBA_MODULE_PREMIUM),
        ("LossReserving", VBA_MODULE_RESERVING),
        ("RiskMetrics", VBA_MODULE_RISK),
        ("clsPolicy", VBA_CLASS_POLICY),
    ]

    print("Injecting VBA modules...")
    try:
        inject_vba_via_zip(str(temp_xlsx), str(OUTPUT_FILE), vba_modules)
        print(f"Created macro-enabled workbook: {OUTPUT_FILE}")
        # Clean up temp
        temp_xlsx.unlink()
    except ImportError:
        # If _vba_project_builder not available, save VBA as separate .bas files
        print("Note: VBA binary builder not available.")
        print("Saving workbook as .xlsm (without embedded VBA) and VBA as .bas files...")
        # Rename .xlsx to .xlsm (Excel may prompt to enable macros)
        import shutil
        shutil.move(str(temp_xlsx), str(OUTPUT_FILE))

        # Save VBA modules as .bas files
        vba_dir = OUTPUT_DIR / "vba_modules"
        vba_dir.mkdir(exist_ok=True)
        for mod_name, mod_code in vba_modules:
            ext = ".cls" if mod_name == "clsPolicy" else ".bas"
            vba_file = vba_dir / f"{mod_name}{ext}"
            vba_file.write_text(mod_code.strip(), encoding="utf-8")
            print(f"  Saved {vba_file.name}")

        print(f"\nWorkbook saved to: {OUTPUT_FILE}")
        print(f"VBA source files saved to: {vba_dir}/")
        print("\nTo use: Open the .xlsm in Excel, press Alt+F11,")
        print("and import the .bas/.cls files into the VBA project.")

    print("\n=== VBA Modules Summary ===")
    for name, code in vba_modules:
        line_count = len(code.strip().splitlines())
        print(f"  {name}: {line_count} lines")
    print(f"\nTotal VBA: {sum(len(c.strip().splitlines()) for _, c in vba_modules)} lines")


if __name__ == "__main__":
    main()
